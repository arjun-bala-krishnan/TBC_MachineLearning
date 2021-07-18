import numpy as np
import pandas as pd
import os
import cv2
import pytesseract
import re
import imutils
import shutil
import xlrd
import math
import json
import datetime
import sys
import time
import exifread
import import_ipynb
import copy
import matplotlib
import h5py

from PIL import Image, ImageEnhance, ImageStat
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from IPython.display import display
from shutil import copyfile, rmtree
from scipy import interpolate
from random import seed, random
from IPython.display import display, HTML
from ipywidgets import IntProgress
from scipy.interpolate import interp1d
from xlrd.sheet import ctype_text 
from scipy.signal import savgol_filter
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import OneHotEncoder
from numpy import argmax

pytesseract.pytesseract.tesseract_cmd = (os.path.join(os.getcwd(), "Tesseract-OCR\\tesseract.exe"))
print("Pandas Version:", pd.__version__)

import tensorflow as tf
import keras
from keras.models import Sequential, Model
from keras.layers import Dense, Reshape, BatchNormalization, MaxPool2D, Convolution2D, InputLayer, Flatten, Concatenate, Lambda, Activation
from keras.utils import plot_model
from sklearn.model_selection import train_test_split
from keras.models import Sequential, Model
from keras import backend as K
print("Tensorflow Version:", tf.__version__)

curr_dir = os.getcwd()  
path_samples = os.path.join(curr_dir, "data")    
path_Reference_Images = os.path.join(path_samples, "#Reference_Images")
path_Post_Process = os.path.join(path_samples, "#Post_Process") 
path_Interfaces = os.path.join(path_samples, "#Interfaces") 
path_Threshold = os.path.join(path_samples, "#Threshold")
path_Roughness = os.path.join(path_samples, "#Roughness")
path_Roughness_roughsurface = os.path.join(path_Roughness, "roughsurface")
path_Roughness_substrate = os.path.join(path_Roughness, "substrate")
path_Models = os.path.join(path_samples, "#Models") 


name_samples = next(os.walk(path_samples))[1]

df_files_data_file = os.path.join(path_samples, "Files_dataframe.json")
df_files_post_process_file = os.path.join(path_samples, "PostProcess_Files_dataframe.json")
df_files_parameters_file = os.path.join(path_samples, "Parameters_dataframe.json")
df_files_roughness_file = os.path.join(path_samples, "Roughness_dataframe.json")


def directory_check():
    if not os.path.isdir(path_samples):
        os.makedirs(path_samples)
    if not os.path.isdir(path_Post_Process):
        os.makedirs(path_Post_Process)     
    if not os.path.isdir(path_Threshold):
        os.makedirs(path_Threshold)
    if not os.path.isdir(path_Interfaces):
        os.makedirs(path_Interfaces)
    if not os.path.isdir(path_Reference_Images):
        os.makedirs(path_Reference_Images)
    if not os.path.isdir(path_Roughness):
        os.makedirs(path_Roughness)
    if not os.path.isdir(path_Roughness_roughsurface):
        os.makedirs(path_Roughness_roughsurface)
    if not os.path.isdir(path_Roughness_substrate):
        os.makedirs(path_Roughness_substrate)
    if not os.path.isdir(path_Models):
        os.makedirs(path_Models)

    print("All Directories Initialized Succesfully")
    
    
def check_empty_json(standard_json):
    if not os.path.isfile(standard_json):
        print(standard_json + " does not exist")
        return(0)
    else:
        print("previous copy of ", standard_json, " exists")
        return(1)
    
def print_full(x):
    pd.set_option('display.max_rows', len(x))
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 2000)
    pd.set_option('display.float_format', '{:20,.2f}'.format)
    pd.set_option('display.max_colwidth', None)
    print(x)
    pd.reset_option('display.max_rows')
    pd.reset_option('display.max_columns')
    pd.reset_option('display.width')
    pd.reset_option('display.float_format')
    pd.reset_option('display.max_colwidth')
    
    
def print_dataframe_full(df):
    with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
        display(df)
    
def view1_image(imageA):
    fig = plt.figure(figsize=(18, 16), dpi= 80, facecolor='w', edgecolor='k')
    ax = fig.add_subplot(1, 2, 1)
    plt.imshow(imageA, cmap = plt.cm.gray)
    plt.axis("off")
    plt.show()

def view2_image(imageA, imageB):
    fig = plt.figure(figsize=(18, 16), dpi= 80, facecolor='w', edgecolor='k')
    ax = fig.add_subplot(1, 2, 1)
    plt.imshow(imageA, cmap = plt.cm.gray)
    plt.axis("off")
    ax = fig.add_subplot(1, 2, 2)
    plt.imshow(imageB, cmap = plt.cm.gray)
    plt.axis("off")
    plt.show()
    
    
def convert_path_to_basename(df_files):
    df_files['Image_Location_Interface'] = df_files['Image_Location_Interface'].apply(lambda x:os.path.basename(x))
    df_files['Image_Location_Threshold'] = df_files['Image_Location_Threshold'].apply(lambda x:os.path.basename(x))
    df_files['Image_Location_PostPocess'] = df_files['Image_Location_PostPocess'].apply(lambda x:os.path.basename(x))    
    return df_files
    
    
def ocr(filename):
    img = cv2.imread(filename, cv2.IMREAD_UNCHANGED)
    height = img.shape[0]
    width = img.shape[1]
    microscope = "Other"
    OCR_Check_Complete = False
    
    ocr_phenom = img[int(0.95 * img.shape[0]):int(0.98 * img.shape[0]), int(0.68 * img.shape[1]):int(0.80 * img.shape[1])]
    ocr_lowres = img[int(0.93 * img.shape[0]):int(0.96 * img.shape[0]), int(0.72 * img.shape[1]):int(0.80 * img.shape[1])]
    
    list_500 = []
    list_2000 = []
    list_500_phenom = [str(i) for i in range(450, 550)] 
    list_2000_phenom = [str(i) for i in range(1950, 2050)] 
    list_500_lowres = [str(i) for i in range(18, 22)]
    list_2000_lowres = [str(i) for i in range(8, 12)]
    
    if height > 2000:
        ocr = ocr_phenom
        list_500 = list_500_phenom
        list_2000 = list_2000_phenom
        microscope = "Phenom"
        
    else:
        ocr = ocr_lowres
        list_500 = list_500_lowres
        list_2000 = list_2000_lowres
        microscope = "LowRes"

    ocr_text = pytesseract.image_to_string(ocr, config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789')
    
    if not ocr_text:
        ocr_text = "NONE"
    else: 
        ocr_text = str((re.findall('\d+', ocr_text))[0])
    
    if ocr_text in list_500:
        ocr_text = 500
        OCR_Check_Complete = True
        print("OCR Check complete = ", ocr_text)
        
    if ocr_text in list_2000:
        ocr_text = 2000   
        OCR_Check_Complete = True
        print("OCR Check complete = ", ocr_text)  
        
    if OCR_Check_Complete is False:
        print("OCR ERROR, Detected text = ", ocr_text)
        print("image location: ", filename)
        view1_image(ocr)
        ocr_text = input("***Enter the magnification manually : ")
        
    return ocr_text




def df_files_update(df_files, start_ID = "JSC100", status = -1):   
    Sample_col = []
    Image_ID_col = []
    Image_Name_col = []
    Image_Type_col = []
    Magnification_col = []  
    Image_Location_col = []
    
    Deleted_File_Locations = []
    Added_File_Locations = []
    
    Sample_col_old = []
    Image_Name_col_old = []
    Image_Type_col_old = []                  
    Image_Location_col_old = []
    Magnification_col_old = []
    
    check = check_empty_json(df_files_data_file)
    
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.splitext(name)[1] == ".tif":
                possible_sample_name_1 = os.path.basename(os.path.dirname(os.path.dirname(os.path.dirname(filename))))
                possible_sample_name_2 = os.path.basename(os.path.dirname(os.path.dirname(filename)))

                if possible_sample_name_1 in name_samples or possible_sample_name_2 in name_samples:                    
                    image_name = os.path.basename(filename)
                    image_location = filename
                    if check ==0:
                        magnification = ocr(filename)

                    if possible_sample_name_1 in name_samples:
                        possible_sample_name = possible_sample_name_1
                        image_type_1 = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                        image_type_2 = os.path.basename(os.path.dirname(filename))
                        image_type = image_type_1 + "-" + image_type_2

                    else:
                        possible_sample_name = possible_sample_name_2
                        image_type = os.path.basename(os.path.dirname(filename))

                    Sample_col.append(possible_sample_name)
                    Image_Name_col.append(image_name)
                    if check ==0:
                        Magnification_col.append(magnification)
                    Image_Type_col.append(image_type)                    
                    Image_Location_col.append(image_location)
                      
          
    if check == 0: 
        df_files['Sample'] = pd.Series(Sample_col)
        df_files['Image_Name'] = pd.Series(Image_Name_col)
        df_files['Image_Type'] = pd.Series(Image_Type_col)
        df_files['Magnification'] = pd.Series(Magnification_col)
        df_files['Image_Location'] = pd.Series(Image_Location_col)
        print("(Status-0) OCR Detection complete..")
        df_files.to_json(df_files_data_file, orient='columns')
        
    else:
        df_files = pd.read_json(df_files_data_file, orient = 'columns')
        df_files.sort_index(inplace=True)
        Sample_col_old = df_files['Sample'].astype(str).tolist()
        Image_Name_col_old = df_files['Image_Name'].astype(str).tolist() 
        Image_Type_col_old = df_files['Image_Type'].astype(str).tolist()  
        Magnification_col_old = df_files['Magnification'].astype(int).tolist() 
        Image_Location_col_old = df_files['Image_Location'].astype(str).tolist()

        Deleted_File_Locations = [item for item in Image_Location_col_old if item not in Image_Location_col]
        Added_File_Locations = [item for item in Image_Location_col if item not in Image_Location_col_old]
        
        if len(Added_File_Locations) > 0:
            for filename in Added_File_Locations:
                image_name = os.path.basename(filename)
                image_location = filename
                magnification = ocr(filename)
                
                possible_sample_name_1 = os.path.basename(os.path.dirname(os.path.dirname(os.path.dirname(filename))))
                possible_sample_name_2 = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                
                if possible_sample_name_1 in name_samples:
                    possible_sample_name = possible_sample_name_1
                    image_type_1 = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                    image_type_2 = os.path.basename(os.path.dirname(filename))
                    image_type = image_type_1 + "-" + image_type_2
                else:
                    possible_sample_name = possible_sample_name_2
                    image_type = os.path.basename(os.path.dirname(filename))   
                
                Sample_col_old.append(possible_sample_name)
                Image_Name_col_old.append(image_name)
                Image_Type_col_old.append(image_type)
                Image_Location_col_old.append(image_location)
                Magnification_col_old.append(magnification)
                
            df_files = pd.DataFrame({'Sample': [],
                'Image_Name': [],
                'Image_ID':[],         
                'Image_Type':[],
                'Magnification': [],
                'Image_Location': [],
                'Image_Location_PostPocess': [],
                'Image_Location_Interface': [],
                'Image_Location_Threshold': []})
            
            df_files['Sample'] = pd.Series(Sample_col_old)
            df_files['Image_Name'] = pd.Series(Image_Name_col_old)
            df_files['Image_Type'] = pd.Series(Image_Type_col_old)
            df_files['Image_Location'] = pd.Series(Image_Location_col_old)
            df_files['Magnification'] = pd.Series(Magnification_col_old)
            print("(Status-1) OCR Detection complete..")
            df_files.to_json(df_files_data_file, orient='columns')
            
            
    if len(Deleted_File_Locations) > 0:
        for filename in Deleted_File_Locations:
            image_name = os.path.basename(filename)
            df_files = df_files[df_files.Image_Name != image_name]
            df_files.to_json(df_files_data_file, orient='columns')
            
    
    df_files = df_files.sort_values(by=['Image_Location'])
    
    ID_No = 0
    for row in df_files.index: 
        Image_ID = "JSC" + str(100 + ID_No)
        ID_No = ID_No + 1
        Image_ID_col.append(Image_ID)
    
    df_files['Image_ID'] = pd.Series(Image_ID_col)
    df_files = df_files.reset_index(drop=True)
    return df_files


def update_reference_column(df_files):
    list_dict = []
    for index, row in df_files.iterrows():
        filename = row['Image_Location']
        directory = os.path.dirname(filename) 
        list_dict.append(directory)

    list_dict = set(list_dict) 
    dict_image_dir = dict.fromkeys(list_dict, [])

    for key_dir in dict_image_dir.keys(): 
        mag_500_status = True
        mag_2000_status = True
        loop_end = False
        for r, d, f in os.walk(key_dir):
            for file in f:
                if '.tif' in file and loop_end is False:
                    filename = os.path.join(r, file)
                    result = df_files['Image_Location'].where(df_files['Image_Location'] == filename)
                    idx = result.tolist().index(filename)
                    mag = df_files.at[idx, 'Magnification']          
                    if mag == 500 and mag_500_status is True:
                        filename_500 = "500_"+ filename
                        mag_500_status = False
                    if mag == 2000 and mag_2000_status is True: 
                        filename_2000 = "2000_"+ filename
                        mag_2000_status = False             

        dict_image_dir[key_dir] = [filename_500, filename_2000]

    def create_ref_img(df):
        filename = df["Image_Location"]
        mag = df["Magnification"]    
        directory = os.path.dirname(filename) 
        if directory in dict_image_dir.keys():
            if str(mag) in dict_image_dir[directory][0][0:4]:
                new_name = dict_image_dir[directory][0][4:]

            if str(mag) in dict_image_dir[directory][1][0:4]:
                new_name = dict_image_dir[directory][1][5:]
        else:
            new_name = "N.A"
        return new_name

    df_files["Reference_Image"] = df_files.apply(create_ref_img, axis = 1)
    
    return df_files


def add_CLAHE_reference_images(df_files_old):
    Ref_Image_List = df_files_old.Reference_Image.unique()
    for img_loc in Ref_Image_List:
        ref_img_pre = cv2.imread(img_loc, cv2.IMREAD_UNCHANGED)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        ref_img_pre_clahe = clahe.apply(ref_img_pre)
        ref_img_pre_clahe = Image.fromarray(ref_img_pre_clahe)
        ref_img_pre_loc = os.path.join(path_Reference_Images, os.path.basename(img_loc))
        ref_img_pre_clahe.save(ref_img_pre_loc)


    def create_changed_ref_img(df):
        ref_location = df["Reference_Image"]
        changed_loc = os.path.join(path_Reference_Images,os.path.basename(ref_location))
        return changed_loc

    df_files_old["Reference_CLAHE"] = df_files_old.apply(create_changed_ref_img, axis = 1)
    return df_files_old


def brightness_CLAHE_eq(df):
    def histogram_pre_process(ref_img_location, img_loc):    
        ref_img = cv2.imread(ref_img_location, cv2.IMREAD_UNCHANGED)
        img = cv2.imread(img_loc, cv2.IMREAD_UNCHANGED)
        ref_img_stat = int(ImageStat.Stat(Image.fromarray(ref_img)).mean[0])
        img_trial = Image.fromarray(img)
        ref_trial_stat = int(ImageStat.Stat(img_trial).mean[0])
        enhanced_img_stat = ref_trial_stat 

        while enhanced_img_stat is not ref_img_stat:
            if enhanced_img_stat < ref_img_stat:  
                enhancer = ImageEnhance.Brightness(img_trial)
                enhanced_img = enhancer.enhance(1.1)
                enhanced_img_stat = ImageStat.Stat(enhanced_img)
                enhanced_img_stat = int(enhanced_img_stat.mean[0])
                img_trial = enhanced_img

            if enhanced_img_stat > ref_img_stat:  
                enhancer = ImageEnhance.Brightness(img_trial)
                enhanced_img = enhancer.enhance(0.99)
                enhanced_img_stat = ImageStat.Stat(enhanced_img)
                enhanced_img_stat = int(enhanced_img_stat.mean[0])
                img_trial = enhanced_img

            if abs(enhanced_img_stat - ref_img_stat) < 3:  
                return img_trial

        return img_trial
    
    
    filename = df["Image_Location"]
    ref = df["Reference_CLAHE"] 
    
    for file, ref_file in zip(filename, ref):
        pre_process_loc = os.path.join(path_Post_Process,os.path.basename(file))
        brighntess_eq_img = histogram_pre_process(ref_file, file)
        brighntess_eq_img.save(pre_process_loc)
        clahe_img = cv2.imread(pre_process_loc, cv2.IMREAD_UNCHANGED)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))       
        brighntess_eq_img_clahe = clahe.apply(clahe_img)
        brighntess_eq_img_clahe = Image.fromarray(brighntess_eq_img_clahe)
        brighntess_eq_img_clahe.save(pre_process_loc)
        
        
def create_pre_process_img(df):
    filename = df["Image_Location"]
    changed_loc = os.path.join(path_Post_Process,os.path.basename(filename))
    return changed_loc


def Otsu_threshold(df):
    Image_Location_PostPocess = df["Image_Location_PostPocess"]
    for file in Image_Location_PostPocess:
        threshold_loc = os.path.join(path_Threshold, os.path.basename(file))
        threshold_img = cv2.imread(file, cv2.IMREAD_UNCHANGED)
        ret, th = cv2.threshold(threshold_img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        th = Image.fromarray(th)
        th.save(threshold_loc)
        
        
def create_th_img(df):
    filename = df["Image_Location"]
    changed_loc = os.path.join(path_Threshold,os.path.basename(filename))
    return changed_loc


def interface(df):
    Image_Location_PostPocess = df["Image_Location_PostPocess"]
    Magnification = df["Magnification"]
    
    for mag, file in zip(Magnification, Image_Location_PostPocess):
        if mag == 500:
            interface_loc = os.path.join(path_Interfaces, os.path.basename(file))
            img = cv2.imread(file, cv2.IMREAD_UNCHANGED)
            roi = img[0:int(0.30 * img.shape[0]), 0:int(img.shape[1])]
            rows = roi.shape[0]
            cols = roi.shape[1]
            blur = cv2.GaussianBlur(roi, (5, 5), 0)
            ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            interface = cv2.Canny(th, 100, 200)
            inter = np.zeros([rows, cols], dtype=int)

            for i in range(0, rows):
                for j in range(0, cols):
                    if interface[i, j] == 255:
                        inter = interface[0:i + 1, j:j + 1].ravel()
                        inter_mean = np.mean(inter[0:i])

                        if inter_mean < 0.1:
                            interface[i, j] = 250

            interface[interface != 250] = 0
            interface[interface == 250] = 255
            cv2.imwrite(interface_loc, interface)

            
def create_inter_img(df):
    filename = df["Image_Location"]
    changed_loc = os.path.join(path_Interfaces,os.path.basename(filename))
    return changed_loc



# delete non existant interface files
def delete_nonexistant_interfaces(Image_Location_Interface):
    if pd.isnull(Image_Location_Interface):
        return np.nan
    if os.path.isfile(Image_Location_Interface):
        return Image_Location_Interface
    else:
        return np.nan


    
def combined_df_files_operations(df_files):
    # Files and image data extraction
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to start Files and image data extraction..")
    else:
        df_files = df_files_update(df_files)

    # The purpose of the following section is to add a column to our dataframe which is the image's reference image
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to add a column of the image's reference image..")
    else:
        df_files_with_reference = update_reference_column(df_files)

    # The purpose here is to apply CLAHE on all reference images and copy them to a separate folder called #Reference_Images, then add to the dataframe
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to apply CLAHE on all reference images and copy them to a separate folder called #Reference_Images..")
    else:
        df_files_with_reference_changed = add_CLAHE_reference_images(df_files_with_reference)

    # The following saves the postprocess image with the same name as filename in a post_process folder doing brightness equilisation after comparison with CLAHE reference image
    # The purpose of this function is to brightness equalise the images based on a reference image from the same subset
    # A part of it was deleted: which was using a linear row strip on the top-most layer to find out the difference from black value and apply it uniformly throughout the image. The plan was dropped because there are certain 500X images which has white matrix crossing over the top boundary

    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to brightness equalise the images based on a reference image from the same subset..")
    else:
        brightness_CLAHE_eq(df_files_with_reference_changed)

    # The operation is to change filenames of all postprocess files with root dir as #Post_Process    
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to change filenames of all postprocess files with root dir as #Post_Process..")
    else:
        df_files_with_reference_changed["Image_Location_PostPocess"] =  df_files_with_reference_changed.apply(create_pre_process_img, axis = 1)

    # We apply simple Otsu threshold to the post processed files
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to apply simple Otsu threshold to the post processed files..")
    else:
        Otsu_threshold(df_files_with_reference_changed)

    # The operation is to change filenames of all threshold files with root dir as #Threshold
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to change filenames of all threshold files with root dir as #Threshold..")
    else:
        df_files_with_reference_changed["Image_Location_Threshold"] =  df_files_with_reference_changed.apply(create_th_img, axis = 1) 

    # Here we isolate the interfaces of the top portion from the post processed image after blur->Canny->light_shining_algorithm from top to determine top pixels
    # Shadow scalares are not calculated yet
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to isolate the interfaces of the top portion..")
    else:
        interface(df_files_with_reference_changed)

    # The operation is to change filenames of all interface files with root dir as #Interfaces    
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to change filenames of all interface files with root dir as #Interfaces ..")
    else:
        df_files_with_reference_changed["Image_Location_Interface"] =  df_files_with_reference_changed.apply(create_inter_img, axis = 1)

    # We get the Interface boundary from top and bottom portion of each microstructure. This is to ensure that slices taken later are within the ROI and also significantly reduce the calculation on custom generator function   
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " get the Interface boundary from top and bottom portion of each microstructure..")
    else:
        df_files = df_files_with_reference_changed.copy()
        df_files['Top_Boundary'] = df_files.apply(lambda row:get_interface_boundaries_top(row['Image_Location_PostPocess'],row['Magnification']),axis=1)
        df_files['Bottom_Boundary'] = df_files.apply(lambda row:get_interface_boundaries_bottom(row['Image_Location_PostPocess'],row['Magnification']),axis=1)

    # The following operation is to covert all non-existent interface files to nan 
    if (os.path.isfile(df_files_post_process_file)):
        print("Delete ", df_files_post_process_file, " to covert all non-existent interface files to NaN..")
    else:
        df_files['Image_Location_Interface'] = df_files.apply(lambda x: delete_nonexistant_interfaces(x['Image_Location_Interface']),axis=1)

    # This saves the completed files as a JSON file for future backup
    if (os.path.isfile(df_files_post_process_file)):
        # Read the saved file
        df_files = pd.read_json(df_files_post_process_file, orient = 'columns') 
        df_files.sort_index(inplace=True)
        print("Delete ", df_files_post_process_file, " to save the completed files as a JSON file..")
    else:
        df_files.to_json(df_files_post_process_file, orient='columns')
        
    return df_files    
    
    
def df_parameters_data_extract():
    parameters_filename_dict = {}
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "parameters.xlsx":
                key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                parameters_filename_dict[key] = filename
    
    row = 0
    parameters_list_dict = {}
    for key in parameters_filename_dict:
        parameters_dict = {}
        xl_workbook = xlrd.open_workbook(parameters_filename_dict[key])
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        num_cols = xl_sheet.ncols   # Number of columns
        num_rows = xl_sheet.nrows
        
        for col_idx in range(0, num_cols):
            cell_key = str(xl_sheet.cell(0, col_idx).value).replace(" ", "")
            cell_value = str(xl_sheet.cell(1, col_idx).value).replace(" ", "")
            parameters_dict[cell_key] = cell_value  
            
        parameters_list_dict[row] = parameters_dict
        row = row + 1
        
    df_parameters = pd.DataFrame.from_dict(parameters_list_dict,orient='index')
    return df_parameters


def hot_encode_normalise(df_parameters, max_powder_dia = 200, div = 200):
    df_copy = copy.deepcopy(df_parameters)
    df_copy = df_copy.apply(df_parameters_bounded(df_copy, max_powder_dia = max_powder_dia, div = div), axis = 1)

    df_copy['Powder_Density_(g/cc)'] = df_copy['Powder_Density_(g/cc)']/(df_copy['Powder_Density_(g/cc)'].max(skipna=True) + 1)
    df_copy['Robot_Arm_Sweep_Velocity_(mm/sec)'] = df_copy['Robot_Arm_Sweep_Velocity_(mm/sec)']/(df_copy['Robot_Arm_Sweep_Velocity_(mm/sec)'].max(skipna=True) + 1)
    df_copy['Spray_Distance_(mm)'] = df_copy['Spray_Distance_(mm)']/(df_copy['Spray_Distance_(mm)'].max(skipna=True) + 1)
    df_copy['Thickness_(microns)'] = df_copy['Thickness_(microns)']/(df_copy['Thickness_(microns)'].max(skipna=True) + 1)
    df_copy['Weight_(g)'] = df_copy['Weight_(g)']/(df_copy['Weight_(g)'].max(skipna=True) + 1)
    df_copy['Coating_Temperature_(Celsius)'] = df_copy['Coating_Temperature_(Celsius)']/(df_copy['Coating_Temperature_(Celsius)'].max(skipna=True) + 1)
    df_copy['Current_(A)'] = df_copy['Current_(A)']/(df_copy['Current_(A)'].max(skipna=True) + 1)

    cat_columns = ["Powder_Name", "Type_of_Coating","Powder_Size_Class", "Powder_Process", "Powder_Shape"]
    df_parameters_encoded = pd.get_dummies(df_copy, prefix_sep="__", columns=cat_columns)

    label_encoders = {}
    for col in cat_columns:
        print("HotEncoding {}".format(col))
        new_le = LabelEncoder()
        df_parameters_encoded[col] = new_le.fit_transform(df_copy[col])
        label_encoders[col] = new_le

    cat_columns_idx = [df_parameters_encoded.columns.get_loc(col) for col in cat_columns]
    ohe = OneHotEncoder()
    df_parameters_encoded_np = ohe.fit_transform(df_parameters_encoded)
    return df_parameters_encoded

    # Example of one hot encode
    # data = np.asarray(df_parameters_encoded['Powder_Name'])
    # label_encoder = LabelEncoder()
    # integer_encoded = label_encoder.fit_transform(data)
    # onehot_encoder = OneHotEncoder(sparse=False)
    # integer_encoded = integer_encoded.reshape(len(integer_encoded), 1)
    # onehot_encoded = onehot_encoder.fit_transform(integer_encoded)
    # print("onehot_encoded = ",onehot_encoded)
    # inverted = label_encoder.inverse_transform([argmax(onehot_encoded[0, :])])
    # print("inverted = ",inverted)


def combined_df_parameters_operations(max_powder_dia = 200, div = 200):
    # Extracting all powder and coating parameters
    if (os.path.isfile(df_files_parameters_file)):
        # Read the saved file
        df_parameters = pd.read_json(df_files_parameters_file, orient = 'columns') 
        df_parameters.sort_index(inplace=True)
        # Based on a bounded class taking maximum powder size (max_powder_dia) and splitting into the three ranges (between 0 and div) corresponding to the size classes
        # deepcopy to get totally independent dataframes
        df_parameters_encoded = hot_encode_normalise(df_parameters, max_powder_dia = max_powder_dia, div = div)
        df_parameters_encoded = df_parameters_encoded.drop(columns = "Powder_Commercial_Name")
        print("Delete ", df_files_parameters_file, " to extract all powder and coating parameters and to save new..")
    else:
        df_parameters = df_parameters_data_extract()
        df_parameters.to_json(df_files_parameters_file, orient='columns')
        
        df_parameters_encoded = hot_encode_normalise(df_parameters, max_powder_dia = max_powder_dia, div = div)
        df_parameters_encoded = df_parameters_encoded.drop(columns = "Powder_Commercial_Name")
        print("Delete ", df_files_parameters_file, " to extract all powder and coating parameters and to save new..")
        
    return df_parameters, df_parameters_encoded


def df_roughness_data_extract():
    df_roughness = pd.DataFrame({'Sample': str,
                       'Roughness_Type': [],
                       'Lambda_c_(microns)': [],
                       'Ra_(microns)': [],
                       'Rq_(microns)': [],
                       'Rz_(microns)': []})
    
    sample = []
    roughness_type = []
    lambda_c = []
    ra = []
    rq = []
    rz = []
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "roughness.txt":
                sample_name = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                sample_type = os.path.basename(os.path.dirname(filename))
                sample.append(sample_name)
                roughness_type.append(sample_type)
                with open(filename, "r") as f:
                    data = f.readlines()
                    for index, line in enumerate(data):
                        if index == 1:
                            words = line.split(";")
                            for index, value in enumerate(words):
                                if index == 4:
                                    lambda_c.append(value)   
                                if index == 6:
                                    ra.append(value)
                                if index == 9:
                                    rq.append(value)
                                if index == 12:
                                    rz.append(value)
     
    df_roughness['Sample'] = pd.Series(sample)
    df_roughness['Roughness_Type'] = pd.Series(roughness_type)
    df_roughness['Lambda_c_(microns)'] = pd.Series(lambda_c)
    df_roughness['Ra_(microns)'] = pd.Series(ra)
    df_roughness['Rq_(microns)'] = pd.Series(rq)
    df_roughness['Rz_(microns)'] = pd.Series(rz)

    return df_roughness

def get_nearest_neighbours(img, pos_i, pos_j, pix_block_size):
    top_NN_i = []
    top_NN_j = []
    left_NN_i = []
    left_NN_j = []
    bottom_NN_i = []
    bottom_NN_j = []
    right_NN_i = []
    right_NN_j = []
    
    for i in range(pix_block_size+2):
        top_NN_i.append(pos_i - 1)
        left_NN_j.append(pos_j - 1)
        bottom_NN_i.append(pos_i + pix_block_size)
        right_NN_j.append(pos_j + pix_block_size)
         
    for i in range(-1,pix_block_size+1,1):
        top_NN_j.append(pos_j + i)
        left_NN_i.append(pos_i + i)
        bottom_NN_j.append(pos_j + i)
        right_NN_i.append(pos_i + i)
        
    top_NN = []
    left_NN = []
    bottom_NN = []
    right_NN = []
    
    for i in range(pix_block_size+2):
        top_NN.append((top_NN_i[i],top_NN_j[i]))
        left_NN.append((left_NN_i[i],left_NN_j[i]))
        bottom_NN.append((bottom_NN_i[i],bottom_NN_j[i]))
        right_NN.append((right_NN_i[i],right_NN_j[i]))
        
    # convert all tuples into list to make it immutable     
    top_NN = list(top_NN)
    left_NN = list(left_NN)
    bottom_NN = list(bottom_NN)
    right_NN = list(right_NN)   
    
    coords = top_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    top_NN = coords
    
    coords = left_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    left_NN = coords
    
    coords = right_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    right_NN = coords
    
    coords = bottom_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    bottom_NN = coords   
    
    coords = top_NN
    top_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            top_NN.append(coord) 
            
    coords = left_NN
    left_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            left_NN.append(coord) 
            
    coords = right_NN
    right_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            right_NN.append(coord) 
            
    coords = bottom_NN
    bottom_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            bottom_NN.append(coord) 
    
    NN_List = top_NN + left_NN + bottom_NN + right_NN    
    NN_List = list( dict.fromkeys(NN_List) )

    return NN_List


# gets the connected black cluster frominput position
def get_cluster_iteration(img, pos_i, pos_j, found_undercut = False):
    shape_cluster = []
    if img[pos_i, pos_j] == 0:
        found_undercut = True
        neighbour = (pos_i, pos_j)
        shape_cluster.append(neighbour)
    else:
        return found_undercut, shape_cluster

    possible_pixels = get_nearest_neighbours(img, pos_i, pos_j, 1)
    for pixel in possible_pixels:
        if pixel in shape_cluster:
            continue

        if img[pixel[0], pixel[1]] == 0:
            shape_cluster.append(pixel)
            img[pixel[0], pixel[1]] = 1

            neighbour_list = get_nearest_neighbours(img, pixel[0], pixel[1], 1)
            for neighbour in neighbour_list:
                possible_pixels.append(neighbour)                          

    if found_undercut is True:
        return found_undercut, shape_cluster
    else:
        return found_undercut, shape_cluster


# gives the shadow property of a single file including the dataframe (which will be combined for distribution)
def return_shadow_parameters(Image_Location_Interface, Image_Location_PostProcess):
    img = cv2.imread(Image_Location_Interface, cv2.IMREAD_UNCHANGED)
    img_orig = cv2.imread(Image_Location_PostProcess, cv2.IMREAD_UNCHANGED)
    x = []
    y_from_top = []
    y_from_base = []
    height_image = img.shape[0] 
    width_image = img.shape[1]

    for i in range(img.shape[0]):
        for j in range(img.shape[1]):
            if img[i,j] == 255:
                x.append(j)
                y_from_top.append(i)
                i_new = img.shape[0] - i
                y_from_base.append(i_new)


    df_interface = pd.DataFrame({'x_orig':x, 'y_from_base':y_from_base, 'y_from_top':y_from_top})
    df_interface = df_interface.sort_values(by=['x_orig'])
    df_interface = df_interface.reset_index()
    df_interface = df_interface.drop(columns = 'index')
    step_diff = df_interface['y_from_base'].values[:-1] - df_interface['y_from_base'].values[1:]
    df_interface = pd.concat([df_interface, pd.DataFrame({'step_diff':step_diff}).shift(1)], axis=1)
    df_interface = df_interface.replace(np.nan, 0)
    

    # Most of the values are 0 because of the nature of adjacent pixels. this is not of interest when calculating shadow effect
    # Most of the remaining values are small (1 or 2), because of incorrect resolution of the image, or thresholding
    # also we remove very large steps, chances are that they are pixels dust present in the resin
    df_interface.drop(df_interface[abs(df_interface.step_diff) < 4].index, inplace=True)
    df_interface.drop(df_interface[abs(df_interface.step_diff) > 100].index, inplace=True)

    # direction of step can be used to create a unimodal distribution around 0 if we have enough undercuts from many images, but now we will not take it into consideation
    df_interface['direction'] = np.where(df_interface['step_diff']>=0.0, -1, +1)
    df_interface['step_diff'] = df_interface['step_diff'].abs().astype(np.int)

    # positive condition means step is rising and negative means step is falling
    
    df_interface['x1_1'] = df_interface['x_orig']
    df_interface['x1_2'] = df_interface['x_orig'] - df_interface['step_diff']
    df_interface['x2_1'] = df_interface['x_orig'] + df_interface['step_diff']
    df_interface['x2_2'] = df_interface['x_orig']
    
    df_interface['y1_1'] = df_interface['y_from_top']
    df_interface['y1_2'] = df_interface['y_from_top']-df_interface['step_diff']
    df_interface['y2_1'] = df_interface['y_from_top']+df_interface['step_diff']
    df_interface['y2_2'] = df_interface['y_from_top']
    
    df_interface['x1_1'] = np.where(df_interface['x1_1'] < 0, 0, df_interface['x1_1'])
    df_interface['x1_2'] = np.where(df_interface['x1_2'] < 0, 0, df_interface['x1_2'])
    df_interface['x2_1'] = np.where(df_interface['x2_1'] > width_image, width_image - 1, df_interface['x2_1'])
    df_interface['x2_2'] = np.where(df_interface['x2_2'] > width_image, width_image - 1, df_interface['x2_2'])
    
    df_interface['y1_1'] = np.where(df_interface['y1_1'] < 0, 0, df_interface['y1_1'])
    df_interface['y1_2'] = np.where(df_interface['y1_2'] < 0, 0, df_interface['y1_2'])
    df_interface['y2_1'] = np.where(df_interface['y2_1'] > height_image, height_image - 1, df_interface['y2_1'])
    df_interface['y2_2'] = np.where(df_interface['y2_2'] > height_image, height_image - 1, df_interface['y2_2'])
    
    
    df_interface['x_slice_start'] = np.where(df_interface['direction']==1, df_interface['x1_1'], df_interface['x1_2'])
    df_interface['x_slice_end'] = np.where(df_interface['direction']==1, df_interface['x2_1'], df_interface['x2_2'])
    df_interface['y_slice_start'] = np.where(df_interface['direction']==1, df_interface['y1_1'], df_interface['y1_2'] )
    df_interface['y_slice_end'] = np.where(df_interface['direction']==1,df_interface['y2_1'], df_interface['y2_2'])

    
    df_interface = df_interface.drop(columns = 'y1_1') 
    df_interface = df_interface.drop(columns = 'y1_2') 
    df_interface = df_interface.drop(columns = 'y2_1') 
    df_interface = df_interface.drop(columns = 'y2_2') 
    df_interface = df_interface.drop(columns = 'x1_1') 
    df_interface = df_interface.drop(columns = 'x1_2') 
    df_interface = df_interface.drop(columns = 'x2_1') 
    df_interface = df_interface.drop(columns = 'x2_2') 
    
    df_interface = df_interface.reset_index(drop=True)
    
    count_undercut = []    
    for idx in range(df_interface.shape[0]):
        x1 = df_interface['x_slice_start'].iloc[idx]
        y1 = df_interface['y_slice_start'].iloc[idx]
        x2 = df_interface['x_slice_end'].iloc[idx]
        y2 = df_interface['y_slice_end'].iloc[idx]
        direction = df_interface['direction'].iloc[idx]
        
        slice_img = img_orig[y1:y2, x1:x2]
        
        _, slice_img_th = cv2.threshold(slice_img, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
        if direction <0:
#             print("first black to left (starting from Right Top) is undercut, extreme left are bad")
            pos_x_const = slice_img_th.shape[1]-1
            pos_y_start = 0
            pos_y_end = slice_img_th.shape[0]-1                

            for pos_y in range(pos_y_end):
                found_undercut, shape_cluster = get_cluster_iteration(slice_img_th, pos_y, pos_x_const, found_undercut = False)
                if found_undercut is True:
                    count = len(shape_cluster)
                    count_undercut.append(count)
                    break
        else:
#             print("first black to right (starting from left Bottom) is undercut, extreme right blacks are bad")
            pos_x_const = 0
            pos_y_start = slice_img_th.shape[0]-1 
            pos_y_end = -1               

            for pos_y in range(pos_y_start, pos_y_end, -1):
                found_undercut, shape_cluster = get_cluster_iteration(slice_img_th, pos_y, pos_x_const, found_undercut = False)
                if found_undercut is True:
                    count = len(shape_cluster)
                    count_undercut.append(count)
                    break

    df_interface = pd.concat([df_interface, pd.DataFrame({'undercut_pixels':count_undercut})], axis=1)
    df_interface = df_interface.replace(np.nan, 0)
    df_interface['undercut_pixels'] = df_interface['undercut_pixels'].astype(np.int)
    # we ignore all steps with undercut that is zero which are essnetially straight lines evident from interface
    df_interface.drop(df_interface[df_interface.undercut_pixels == 0].index, inplace=True)
    df_interface = df_interface.reset_index()
    df_interface = df_interface.drop(columns = 'index') 

    count = 0
    step_sum = 0
    undercuts_sum = df_interface['step_diff'].sum()
    count_undercuts = df_interface['step_diff'].shape[0]
    

    shadow_average = undercuts_sum/count_undercuts
    shadow_density = count_undercuts/width_image     
        
    total_undercut_area = df_interface['undercut_pixels'].sum()
    total_undercut_possible_area = (df_interface['step_diff']**2).sum()
    undercuts_area_average = total_undercut_area/total_undercut_possible_area

    return df_interface, shadow_average, shadow_density, undercuts_area_average



# we will find the shadow property for the whole set of one sample and evaluate the fitness with each distribution in the d10 and d90 parts
def make_interface_shadow_parameters(df_roughness):
    samples = df_roughness['Sample'].unique()
    df_master_interface_sample = dict()
    dict_interface_parameteters = {'Image_Location_Interface': [], 'shadow_average': [], 'shadow_density': [], 'undercuts_area_average': []}
    keys_interface_parameteters = ['Image_Location_Interface', 'shadow_average', 'shadow_density', 'undercuts_area_average']
    
    for sample in samples:
        filter1 = df_roughness['Sample'] == sample
        df_sample_interface = df_roughness[filter1][['Image_Location_Interface']]
        df_sample_postprocess = df_roughness[filter1][['Image_Location_PostPocess']]
        print("Started collecting data of: ", sample)
        list_of_dataframes= []
        for interface_file_path, postprocess_file_path in zip(df_sample_interface.values, df_sample_postprocess.values):
            Image_Location_Interface = interface_file_path[0]
            print("Working on: ", Image_Location_Interface)
            df_interface_component, shadow_average, shadow_density, undercuts_area_average = return_shadow_parameters(Image_Location_Interface, postprocess_file_path[0])
            print("shadow_average = ", shadow_average," shadow_density = ", shadow_density, "undercuts_area_average = ", undercuts_area_average)                        
            list_of_dataframes.append(df_interface_component) 
            for key in keys_interface_parameteters:
                dict_interface_parameteters[key].append(eval(key))   
            
        df_master = pd.concat(list_of_dataframes)
        df_master_interface_sample[sample] = df_master                               
            
    return df_master_interface_sample, dict_interface_parameteters




def generate_roughness_slices(csv_file, sq_diam = 224, kernel = 5, surface = "Rough", image = True, encoded = True, color_coded = True, seed = None):
    output_roughness_image = np.zeros((sq_diam, sq_diam), np.uint8)
    slices_encoded_dict = dict()
    slices_color_coded_dict = dict()
    
    
    if seed is not None:
        np.random.seed(seed) 
        
    df = pd.read_csv(csv_file) 
    df = df.drop(columns = "Unnamed: 0")
    random_row = np.random.randint(0, df.shape[0]- sq_diam - 1)
    output_roughness_orig_file = np.nan
    
    # this generates an image file using plot function with a resolution according to the dpi and figure size provided
    if image is True:
        canvas = np.zeros((sq_diam, sq_diam), np.uint8)
        # x is already in microns
        x = df['x'][random_row:random_row+224].values
        z = ((df['z'][random_row:random_row+224])*1000).values
        
        x = x - np.min(x)
        z = z - np.min(z)
        each_pix = np.max(z)/(sq_diam-1)
        new_z = z/each_pix
        new_z = new_z.astype(int)
        new_z = abs(np.subtract(new_z, 223))
        xz = list(zip(new_z, x))
        
            
        for idx, coord in enumerate(xz):
            if idx>0:
                old_coord = xz[idx-1]
                new_coord = xz[idx]
                cv2.line(canvas, (old_coord[1],old_coord[0]), (new_coord[1],new_coord[0]), 255)

        output_roughness_image = canvas
    
    # generates a dictionary of scalar list
    if encoded is True:
        vals = df['z'].values
        y_diffs = vals[:-1] - vals[1:]
        angles = np.arctan(y_diffs*1000)
        angles = np.degrees(angles)
        df = pd.concat([df, pd.DataFrame({'y_diff':y_diffs})], axis=1) 
        df = pd.concat([df, pd.DataFrame({'angles':angles})], axis=1) 
        
        max_height = df['y_diff'][random_row:random_row+sq_diam].max() 
        min_height = df['y_diff'][random_row:random_row+sq_diam].min()
        max_angle = 90.0
        min_angle = 0.0
        
        df = df[random_row:random_row+224]
        df.reset_index(inplace=True)
        df = df.drop(columns=['index'])
        keys = ['y_diff_arr', 'angles_arr', 'intensity_value_arr', 'direction_value_arr']
        
        y_diff_arr  = [] 
        angles_arr = []
        intensity_value_arr = []
        direction_value_arr = []
        
        for pos in range(sq_diam):
            y_diff = df['y_diff'].iloc[pos]
            angles = df['angles'].iloc[pos]
            intensity_value = df['pix'].iloc[pos]

            if y_diff<0:   
                direction_value = 0
            if y_diff>=0:
                direction_value = 1
                
            y_diff = abs(y_diff) 
            angles = abs(angles)
            
            increment = (max_height -  min_height)
            y_diff_pix_value = (y_diff - min_height)/increment
            y_diff_arr.append(y_diff_pix_value)
            angles_arr.append(angles)
            intensity_value_arr.append(intensity_value)
            direction_value_arr.append(direction_value)
        
        y_diff_arr = np.asarray(y_diff_arr, dtype = np.float32)
        angles_arr = np.asarray(angles_arr, dtype = np.float32)/90.0
        intensity_value_arr = np.asarray(intensity_value_arr, dtype = np.float32)/255.0 
        direction_value_arr = np.asarray(direction_value_arr, dtype = np.int8)
        
        for key in keys:
            slices_encoded_dict[key] = eval(key)
            
    
    # this generates a set of RGB images currosponding to the slice (this part has been ignored since 3D convolution produces no more results than normal layers)
    if color_coded is True:
        if kernel == 1:
            no_of_canvas = 2
        else:
            no_of_canvas = int(kernel)
            
        vals = df['z'].values
        y_diffs = vals[:-1] - vals[1:]

        angles = np.arctan(y_diffs*1000)
        df_2 = pd.DataFrame({'y_diff':y_diffs})
        angles = np.degrees(angles)
        df_3 = pd.DataFrame({'angles':angles})
        df = pd.concat([df, df_2], axis=1) 
        df = pd.concat([df, df_3], axis=1) 
        
        # to eliminate the vast difference that occurs when we compare the whole dataset, it is better to get the local maxima and minima for scaling
        # would it better to use some log value to scale the entire array? otherwise the heights will be very localised
        max_height = df['y_diff'][random_row:random_row+sq_diam].max() 
        min_height = df['y_diff'][random_row:random_row+sq_diam].min()
        
        # if we wish to use the angles for scaling by 255 equal parts using color intensity, this is not done now since we use angular displacements to plot
        max_angle = 90.0
        min_angle = 0.0

        # the row position that should be taken from the whole dataset, column always start from 0, local row value changes with pos initializations
        # the number of angles that it can represent depends on kernel size, it has to be distributed among significant_pixels
        # one among the ignificant_pixels represents the angle end point (with resolution depending on kernel size)
        # the choice_pixel is the number chosen with x as base and through to the corener or the currosponding y base calculated as a global y, 
        pos = random_row
        init_pos = pos
        final_pos = init_pos
        diff_pos = 0
        slice_no = 1
        
        while diff_pos < sq_diam:
            canvas = np.zeros((kernel, sq_diam,3), np.uint8)
            col = 0
            row = kernel-1
            significant_pixels = kernel+ (kernel-1)
            choice_pixel = 0

            for i in range(int(sq_diam*kernel - kernel)):
                y_diff = df['y_diff'].iloc[pos]
                angles = df['angles'].iloc[pos]

                # PLOT lines are angles in abs values. since we convert all the angles to 1st quadrant. The direction bit will be implimented in blue color (+/0 = 255 or - = 0)
                # RED = intensity_pix_value
                # Blue = direction
                # Green = "local" y_diff_pix_value

                angles = abs(angles)
                if angles>90:
                    angles = 90

                intensity_pix_value = df['pix'].iloc[pos]

                if y_diff<0:   
                    direction = 0
                if y_diff>=0:
                    direction = 255

                y_diff = abs(y_diff)  
                increment = (max_height -  min_height)/255
                y_diff_pix_value = (y_diff - min_height)/increment
                y_diff_pix_value = y_diff_pix_value.astype(np.uint8)

                # Now we generate a slice currosponding to sqdiam/kernel size
                if i%kernel == 0:
                    if col%kernel == 0:
                        # finding a whole number that represents the boundary pixel within kernal to which the line should be drawn
                        # the choice pixel is then converted in to a global x and y cordinate
                        choice_pixel = int(angles/(90.0/significant_pixels)) 
                        if choice_pixel < int(kernel):
                            choice_pixel_col = col + kernel - 1
                            choice_pixel_row = row - (choice_pixel)
                        else:
                            choice_pixel_col = col + (2*kernel - choice_pixel) - 2
                            choice_pixel_row = row - (kernel - 1)

                        # assigning color with weights found previously    
                        color = (int(direction), int(y_diff_pix_value), int(intensity_pix_value))
                        cv2.line(canvas, (col, row), (choice_pixel_col, choice_pixel_row), tuple (color))
                        final_pos = final_pos + 1
                    
                    row = kernel -1
                    col = col + 1
                    pos = pos + 1
            
            slices_color_coded_dict[str(slice_no)] = canvas
            diff_pos = final_pos - init_pos
            pos = final_pos
            slice_no = slice_no + 1
            
    return output_roughness_image, slices_encoded_dict, slices_color_coded_dict


# if df_roughness.empty or df_files.empty:
#     df_roughness = df_roughness_data_extract()
#     check_roughness_file()
#     df_roughness['scan_csv_location']=df_roughness.apply(lambda x: update_csv_location(x['Sample'],x['Roughness_Type']),axis=1)
#     df_files = pd.read_json(df_files_post_process_file, orient = 'columns') 
#     df_files.sort_index(inplace=True)
#     df_roughness_interface_files_rough_surface, df_roughness_interface_files_substrate = associate_appropriate_scan_files(df_roughness, df_files)
# else:
#     df_roughness_interface_files_rough_surface, df_roughness_interface_files_substrate = associate_appropriate_scan_files(df_roughness, df_files)



def df_mercury_data_extract():
    Hg_dict = {}
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "Hg_Porosimetry.xlsx":
                key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                Hg_dict[key] = filename

    start_row = 0
    Hg_df_dict = {}
    datas_ = []
    for key in Hg_dict:
        xl_workbook = xlrd.open_workbook(Hg_dict[key])
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        num_cols = xl_sheet.ncols   # Number of columns
        num_rows = xl_sheet.nrows
        for row_idx in range(0, num_rows):    # Iterate through rows
            row = row_idx
            for col_idx in range(0, num_cols):  # Iterate through columns
                cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                col = col_idx
                cell_obj_trunc = str(cell_obj.value).replace(" ", "")
                if cell_obj_trunc == "DATAREPORT":
                    start_row = row + 1

        excel_file = Hg_dict[key]
        df = pd.read_excel(excel_file)
        df_sheet1 = pd.read_excel(excel_file, sheet_name = 0, skiprows = start_row)
        df_sheet1 = df_sheet1.drop([0])
        df_sheet1.reset_index(inplace=True)
        df_sheet1 = df_sheet1.drop(columns=['index'])
        df_sheet1 = df_sheet1.loc[:, ['Radius', 'P[%]']]
        df_sheet1.dropna(inplace=True)
        datas_.append((key, df_sheet1['Radius'].values, df_sheet1['P[%]'].values))
        Hg_df_dict[key] = df_sheet1
        
    return datas_, Hg_df_dict



def find_step_distribution_parameters(df_roughness, df_master_interface_sample):
    samples = df_roughness['Sample'].unique()
    df_new = copy.deepcopy(df_master_interface_sample)
    df_final = pd.DataFrame(samples, columns =['Sample']) 
    
    
    dict_total_undercut_ratio = dict()
    dict_p10_steps = dict()
    dict_p50_steps = dict()
    dict_p90_steps = dict()
    dict_p10_undercut_ratio = dict()
    dict_p50_undercut_ratio = dict()
    dict_p90_undercut_ratio = dict()
    
    for sample in samples:
        df = df_new[sample]
        df = df.drop(columns = 'x_orig') 
        df = df.drop(columns = 'y_from_base') 
        df = df.drop(columns = 'y_from_top')
        df = df.drop(columns = 'x_slice_start')
        df = df.drop(columns = 'x_slice_end')
        df = df.drop(columns = 'y_slice_start')
        df = df.drop(columns = 'y_slice_end')
        
        
        df['freq'] = df.groupby('step_diff')['step_diff'].transform('count')
        df['undercut_pixels_ratio'] = df['undercut_pixels']/df['step_diff']**2
        df.drop(df[abs(df.undercut_pixels_ratio) > 0.9].index, inplace=True)
        undercut_pixels_ratio_sum = df['undercut_pixels_ratio'].sum()/df['step_diff'].shape[0]
        
        # total undercuts
        dict_total_undercut_ratio[sample] = undercut_pixels_ratio_sum
        df = df.drop(columns = 'undercut_pixels')
#         df = df.drop_duplicates(subset = 'step_diff')
        df = df.sort_values(by=['step_diff'])
        df['bell_step_diff'] = df['step_diff']*df['direction']
        df = df.reset_index()
        df = df.drop(columns = 'index')
        
        # percentile of steps
        p10_steps = np.percentile(df['step_diff'], 10)
        p50_steps = np.percentile(df['step_diff'], 50)
        p90_steps = np.percentile(df['step_diff'], 90)
        dict_p10_steps[sample] = int(p10_steps)
        dict_p50_steps[sample] = int(p50_steps)
        dict_p90_steps[sample] = int(p90_steps)
        
        # percentile of undercut areas
        p10_undercut_ratio = np.percentile(df['undercut_pixels_ratio'], 10)
        p50_undercut_ratio = np.percentile(df['undercut_pixels_ratio'], 50)
        p90_undercut_ratio = np.percentile(df['undercut_pixels_ratio'], 90)
        dict_p10_undercut_ratio[sample] = p10_undercut_ratio
        dict_p50_undercut_ratio[sample] = p50_undercut_ratio
        dict_p90_undercut_ratio[sample] = p90_undercut_ratio
        
    
    df_final['total_undercut_ratio'] = df_final['Sample'].map(dict_total_undercut_ratio)
    
    df_final['p10_steps'] = df_final['Sample'].map(dict_p10_steps)
    df_final['p50_steps'] = df_final['Sample'].map(dict_p50_steps)
    df_final['p90_steps'] = df_final['Sample'].map(dict_p90_steps)
    df_final['p10_undercut_ratio'] = df_final['Sample'].map(dict_p10_undercut_ratio)
    df_final['p50_undercut_ratio'] = df_final['Sample'].map(dict_p50_undercut_ratio)
    df_final['p90_undercut_ratio'] = df_final['Sample'].map(dict_p90_undercut_ratio)
    
    return df_final





def df_roughness_csv_extract(path_Roughness_folder, roughness_file):    
    roughness_dict = {}
    diam = 1414
    colnames=['x', 'y', 'z', 'global_max', 'global_min', 'pix_value']  
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == roughness_file:
                key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                roughness_dict[key] = filename
    
    global_max = 0.0
    global_min = 1.7976931348623157e+308    # a high value
    
    for key in roughness_dict:
        data = pd.read_csv(roughness_dict[key], names=colnames, header=None)
        # converting units to microns only for pixel brightness calculation   
        data['z'] = data['z']
        
        max_height = data['z'].max()  
        min_height = data['z'].min()
        
        if max_height>global_max:
            global_max = max_height
        if min_height<global_min:
            global_min = min_height
            
    
    for key in roughness_dict:
        data = pd.read_csv(roughness_dict[key], names=colnames, header=None)   
        # converting units to microns
        data['x'] = data['x'] * 1000
        data['y'] = data['y'] * 1000

        
        min_width = data['y'].min()
        min_length = data['x'].min()
        data['x'] = data['x'].subtract(min_length)
        data['y'] = data['y'].subtract(min_width)
        
        data['global_max'].values[:] = global_max
        data['global_min'].values[:] = global_min

        increment = (data['global_max'] -  data['global_min'])/255.0
        data['pix_value'] = (data['z'] - data['global_min'])/increment
        data['pix_value'] = data['pix_value'].astype(np.uint8)
        data['z'] = data['z'] - data['global_min']
        
        data['global_max'] = data['global_max'] - data['global_min']
        data['global_min'] = data['global_min'] - data['global_min']
        
        data_group = data.groupby(['x'])['y', 'z', 'global_max', 'global_min', 'pix_value']
        df_dict = {}
        
        for idx, data_part in data_group:
            data_part['x'] = data_part['x'].astype(np.uint16)
            data_part['y'] = data_part['y'].astype(np.uint16)
            data_part['z'] = data_part['z'].astype(np.float32)
            tot_points = data_part['z'].count()
            data_part['global_max'] = data_part['global_max'].astype(np.uint16)
            data_part['global_min'] = data_part['global_min'].astype(np.uint16)
            img_height = data_part['global_max'].max() + 1

            x_coord_array = []
            z_coord_array = []
            pix_value_array = []

            for x_coord in range(tot_points):          
                z_coord = data_part['z'].iloc[x_coord]
                pix_value = data_part['pix_value'].iloc[x_coord]

                x_coord_array.append(x_coord) 
                z_coord_array.append(z_coord) 
                pix_value_array.append(pix_value) 

            df = pd.DataFrame({'x':x_coord_array, 'z':z_coord_array, 'pix':pix_value_array}) 
            df_dict[str(int(idx))] = df
                
        path_csv_file = os.path.join(path_Roughness_folder, key + ".csv")
        frames = []
        for key in df_dict:
            frames.append(df_dict[key])

        combined_df = pd.concat(frames, ignore_index=True)
        combined_df.to_csv(path_csv_file) 

        
def check_roughness_file():
    if len(os.listdir(path_Roughness_roughsurface)) == 0:
        print(path_Roughness_roughsurface, " is empty...extracting csv files")
        # get all the scanned data in desired format and save the csv files in their respective folders
        df_roughness_csv_extract(path_Roughness_roughsurface, roughness_file = "scan_roughsurface.csv")
    else:    
        print(path_Roughness_roughsurface, " is not empty")

    if len(os.listdir(path_Roughness_substrate)) == 0:
        print(path_Roughness_substrate, " is empty...extracting csv files")
        # get all the scanned data in desired format and save the csv files in their respective folders
        df_roughness_csv_extract(path_Roughness_substrate, roughness_file = "scan_substrate.csv")
    else:    
        print(path_Roughness_substrate, " is not empty")        
        
        
def update_csv_location(sample, roughness_type):
    if roughness_type == 'Roughness_FreeStanding_Rough_Surface':
        file_path = os.path.join(path_Roughness_roughsurface, sample + '.csv')
        if os.path.isfile(file_path):
            return file_path
    if roughness_type == 'Roughness_FreeStanding_Substrate':
        file_path = os.path.join(path_Roughness_substrate, sample + '.csv')
        if os.path.isfile(file_path):
            return file_path
    
    return np.nan        
  
    
def associate_appropriate_scan_files(df_roughness, df_files):
    filter1 = df_roughness['Roughness_Type'] == 'Roughness_FreeStanding_Rough_Surface'
    df_roughness_1 = df_roughness[filter1]
    df_roughness_1.reset_index(inplace=True)
    df_roughness_1 = df_roughness_1.drop(columns=['index'])
    filter2 = df_files['Magnification'] == 500 
    df_files_1 = df_files[filter2][['Sample','Image_Location_Interface', 'Image_Location_PostPocess']] 
    df_files_1.reset_index(inplace=True)
    df_files_1 = df_files_1.drop(columns=['index'])
    
    df_roughness_interface_files_rough_surface = df_files_1[['Sample', 'Image_Location_Interface', 'Image_Location_PostPocess']].merge(df_roughness_1, left_on='Sample', right_on='Sample')
    
    filter1 = df_roughness['Roughness_Type'] == 'Roughness_FreeStanding_Substrate'
    df_roughness_1 = df_roughness[filter1]
    filter2 = df_files['Magnification'] == 500 
    df_files_1 = df_files[filter2][['Sample','Image_Location_Interface', 'Image_Location_PostPocess']]
    
    df_roughness_interface_files_substrate = df_files_1[['Sample', 'Image_Location_Interface', 'Image_Location_PostPocess']].merge(df_roughness_1, left_on='Sample', right_on='Sample')
    
    return df_roughness_interface_files_rough_surface, df_roughness_interface_files_substrate    
    
    
    
    
# This version result in a list of images for each sample. more concise
def associate_appropriate_scan_files_aslist(df_roughness, df_files):
    filter1 = df_roughness['Roughness_Type'] == 'Roughness_FreeStanding_Rough_Surface'
    df_roughness_1 = df_roughness[filter1]
    
    df_roughness_1.reset_index(inplace=True)
    df_roughness_1 = df_roughness_1.drop(columns=['index'])

    print("df_roughness_1")
    display(df_roughness_1)
       
    filter2 = df_files['Magnification'] == 500 
    df_files_1 = df_files[filter2][['Sample','Image_Location_Interface']].groupby('Sample').agg(list).reset_index()
 
    print("df_files_1")
    display(df_files_1)
    print(df_files_1['Image_Location_Interface'][3])
    
    df_roughness_interface_files_rough_surface = df_files_1[['Sample', 'Image_Location_Interface']].merge(df_roughness_1, left_on='Sample', right_on='Sample')
    
    filter1 = df_roughness['Roughness_Type'] == 'Roughness_FreeStanding_Substrate'
    df_roughness_1 = df_roughness[filter1]
    filter2 = df_files['Magnification'] == 500 
    df_files_1 = df_files[filter2][['Sample','Image_Location_Interface']]
    df_roughness_interface_files_substrate = df_files_1[['Sample', 'Image_Location_Interface']].merge(df_roughness_1, left_on='Sample', right_on='Sample')
    
    return df_roughness_interface_files_rough_surface, df_roughness_interface_files_substrate    
    
        

def df_mercury_common_cordinates(datas_, Hg_df_dict, df_files_new, num = 10, kind = "linear"):  
    def reinterpolate(x_old, y_old, x_new, kind = "linear"):
        f = interp1d(x_old, y_old, kind = kind)
        return f(x_new)

    Hg_porosity_df = pd.DataFrame({'Sample': str,
                       'Radii': [],
                       'Ps': [],
                       'log_Radii': [],
                       'log_Radii_new': [],
                       'Ps_new': []})

    Hg_porosity_df_temp = pd.DataFrame({'Sample': str,
                       'Radii': [],
                       'Ps': [],
                       'log_Radii': [],
                       'log_Radii_new': [],
                       'Ps_new': []})
    
    porosity_df = pd.DataFrame(datas_, columns=['Sample',"Radii","Ps"])
    porosity_df["Radii"] = porosity_df["Radii"].map(lambda array: np.array(array).astype(float))
    porosity_df["log_Radii"] = porosity_df["Radii"].map(np.log)
    rlogmin = porosity_df["log_Radii"].map(min).max()
    rlogmax  =porosity_df["log_Radii"].map(max).min()
    log_Radii_new = np.linspace(rlogmin, rlogmax, num = num)

    
    for index, row in df_files_new.iterrows():
        Sample = row['Sample']
        Hg_df = Hg_df_dict[Sample]
        Radii = Hg_df['Radius'].tolist()
        log_Radii = np.log(Radii)
        Ps = Hg_df['P[%]'].tolist()
        Ps_new = reinterpolate(log_Radii, Ps, log_Radii_new, kind = kind)

        Hg_porosity_df_temp['Sample'] = pd.Series(Sample)
        Hg_porosity_df_temp['Radii'] = pd.Series([Radii])
        Hg_porosity_df_temp['Ps'] = pd.Series([Ps])
        Hg_porosity_df_temp['log_Radii'] = pd.Series([log_Radii])
        Hg_porosity_df_temp['log_Radii_new'] = pd.Series([log_Radii_new])
        Hg_porosity_df_temp['Ps_new'] = pd.Series([Ps_new])
        Hg_porosity_df = Hg_porosity_df.append(Hg_porosity_df_temp, ignore_index = True)
        
    return Hg_porosity_df


def df_parameters_bounded(df_parameters_bounded, max_powder_dia, div):
    # scaling
    df_parameters_bounded['Powder_Diameter_0.1_(microns)'] = (df_parameters_bounded['Powder_Diameter_0.1_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
    df_parameters_bounded['Powder_Diameter_0.5_(microns)'] = (df_parameters_bounded['Powder_Diameter_0.5_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
    df_parameters_bounded['Powder_Diameter_0.9_(microns)'] = (df_parameters_bounded['Powder_Diameter_0.9_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
    
    # normalising
    df_parameters_bounded['Powder_Diameter_0.1_(microns)'] = df_parameters_bounded['Powder_Diameter_0.1_(microns)']/div
    df_parameters_bounded['Powder_Diameter_0.5_(microns)'] = df_parameters_bounded['Powder_Diameter_0.5_(microns)']/div
    df_parameters_bounded['Powder_Diameter_0.9_(microns)'] = df_parameters_bounded['Powder_Diameter_0.9_(microns)']/div                                                          
  
    return df_parameters_bounded




def iter_scale_powder_diameter(df_parameters, classification = 0.1):
    measure_YSZ412M = []
    measure_YSZ413H = []
    measure_YSZ479M = []
    measure_YSZ481M = []
    
    if classification == 0.1:
        dia_class = 'Powder_Diameter_0.1_(microns)'
    elif classification == 0.5:
        dia_class = 'Powder_Diameter_0.5_(microns)'
    elif classification == 0.9:
        dia_class = 'Powder_Diameter_0.9_(microns)'
    else:
        print("Wrong powder class..")
        return

    for div in range(200):
        if div == 0:
            continue

        df_parameters_new = copy.deepcopy(df_parameters)
        df_parameters_new = df_parameters_new.apply(df_parameters_bounded(df_parameters_new, max_powder_dia = 200, div = div), axis = 1)
        measure_YSZ412M.append(df_parameters_new[dia_class][0])
        measure_YSZ413H.append(df_parameters_new[dia_class][5])
        measure_YSZ479M.append(df_parameters_new[dia_class][7])
        measure_YSZ481M.append(df_parameters_new[dia_class][11])



    matplotlib.rcParams['text.usetex']=False
    matplotlib.rcParams['text.latex.unicode']=False
    matplotlib.rc('text', usetex = False)
    matplotlib.rc('font', **{'family':"sans-serif"})
    params = {'text.latex.preamble': [r'\usepackage{siunitx}', 
        r'\usepackage{sfmath}', r'\sisetup{detect-family = true}',
        r'\usepackage{amsmath}']}   
    plt.rcParams.update(params)  

    plt.plot(measure_YSZ479M, label = 'YSZ479M $\mu$m')
    plt.plot(measure_YSZ481M, label = 'YSZ481M $\mu$m')
    plt.plot(measure_YSZ412M, label = 'YSZ412M $\mu$m')
    plt.plot(measure_YSZ413H, label = 'YSZ413H $\mu$m')
    plt.legend()
    
    
    
    

    


def get_interface_boundaries_top(img_loc, mag):
    img = cv2.imread(img_loc, cv2.IMREAD_UNCHANGED)
        
    # The 500 magnification images have a top interface (except the low resolution images from second microscope, chance are low)
    if int(mag) == 500:
        horizontal_kernel = np.ones((1,2048), np.uint8)  # note this is a horizontal kernel
        kernel = np.ones((1,20), np.uint8)
        top_row = 0
        bottom_row = 0

        top_section = img[0:int(0.30 * img.shape[0]), 0:int(img.shape[1])]    
        rows = top_section.shape[0]
        cols = top_section.shape[1]
        blur = cv2.GaussianBlur(top_section, (5, 5), 0)
        ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        top_section_interface = cv2.Canny(th, 100, 200)
        inter = np.zeros([rows, cols], dtype=int)
        for i in range(0, rows):
            for j in range(0, cols):
                if top_section_interface[i, j] == 255:
                    inter = top_section_interface[0:i + 1, j:j + 1].ravel()
                    inter_mean = np.mean(inter[0:i])

                    if inter_mean < 0.1:
                        top_section_interface[i, j] = 250

        top_section_interface[top_section_interface != 250] = 0
        top_section_interface[top_section_interface == 250] = 255

        dil_top_section_interface = cv2.dilate(top_section_interface, kernel, iterations=1)
        ero_top_section_interface = cv2.erode(dil_top_section_interface, kernel, iterations=1) 

        count = np.zeros(rows, dtype=int)
        for i in range(0, rows):
            for j in range(0, cols):
                if top_section_interface[i, j] == 255:
                    count[i] = count[i] + 1

        max_count_row = np.argmax(count)
        for i in range(max_count_row, rows):
            if count[i] == 0:
                top_row = i
                break
    else:
        top_row = 0

    print("top_row cordinate of ",os.path.basename(img_loc) ," = ", top_row)
    return top_row



def get_interface_boundaries_bottom(img_loc, mag):
    img = cv2.imread(img_loc, cv2.IMREAD_UNCHANGED)
    bottom_row = 0
    tot_rows = img.shape[0]
    
    if int(mag) == 2000:
        bottom_row = img.shape[0] - 135
        # extra condition for low res image with smaller legend
        if int(tot_rows) < 1000:
            bottom_row = img.shape[0] - 50
          
    else: 
        bottom_section = img[int(0.70 * img.shape[0]):img.shape[0] - 135, 0:int(img.shape[1])]  
        if int(tot_rows) < 1000:
            bottom_section = img[int(0.70 * img.shape[0]):img.shape[0] - 50, 0:int(img.shape[1])] 
            
        rows = bottom_section.shape[0]
        cols = bottom_section.shape[1]

        ret, th1 = cv2.threshold(bottom_section, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        inter = th1[rows - 1:rows, 0:cols].ravel() 
        count = np.sum(inter)/len(inter)
        # If the strip has legend then it is deleted first
        if count > 200:
            bottom_row = img.shape[0] - 135
            # extra condition for low res image with smaller legend
            if int(tot_rows) < 1000:
                bottom_row = img.shape[0] - 50

        else: 
            blur = cv2.GaussianBlur(bottom_section, (5, 5), 0)
            ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            bottom_section_interface = cv2.Canny(th, 100, 200)
            inter = np.zeros([rows, cols], dtype=int)
            # image without legend is scanned from top after threshold , the canny operator brings grains all across matrix
            for i in range(0, rows):
                for j in range(0, cols):
                    if bottom_section_interface[i, j] == 255:
                        inter = bottom_section_interface[i:rows, j:j + 1].ravel()           
                        inter_count = np.count_nonzero(inter[0:rows-i])
                        if inter_count <= 1:
                            bottom_section_interface[i, j] = 250
            
            bottom_section_interface[bottom_section_interface != 250] = 0
            bottom_section_interface[bottom_section_interface == 250] = 255   
            count = np.zeros(rows, dtype=int)
            for i in range(0, rows):
                for j in range(0, cols):    
                    if bottom_section_interface[i, j] == 255:
                        count[i] = count[i] + 1   

            max_count_row = np.argmax(count)
            for i in range(200, max_count_row):
                if count[i] > 5:
                    bottom_row = i
                    break

            bottom_row = rows - bottom_row
            bottom_row = img.shape[0] - bottom_row - 135
            
            # extra condition for low res image with smaller legend
            if int(tot_rows) < 1000:
                bottom_row = img.shape[0] - bottom_row - 50
    
    print("bottom_row cordinate of ",os.path.basename(img_loc) ," = ", bottom_row,"\n")
    return bottom_row








# Below is a command to find adaptive threshold from openCV implimentation
# Image_Location_PostPocess = df_files_with_reference_changed["Image_Location"]
# dst_newfolder = "C:\\Users\\arjun\\Downloads"
# for i in range(3):
#     img = Image_Location_PostPocess[i]
#     img_new = cv2.imread(img, cv2.IMREAD_UNCHANGED)
#     dst_loc = os.path.join(dst_newfolder, os.path.basename(Image_Location_PostPocess[i]))
#     dst = cv2.adaptiveThreshold(img_new, maxValue = 255, adaptiveMethod = cv2.BORDER_REPLICATE, thresholdType = cv2.THRESH_BINARY, blockSize=151, C = 5)
#     view2_image(img_new, dst)
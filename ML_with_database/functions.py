import numpy as np
import os
import pandas as pd
import cv2
import pytesseract
import re
import imutils
import shutil
import xlrd
import h5py
import math
import json
import datetime
import sys
import time
import keras

from PIL import Image, ImageEnhance, ImageStat
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from IPython.display import display
from shutil import copyfile, rmtree
from scipy import interpolate
from random import seed, random
from IPython.display import display, HTML
from ipywidgets import IntProgress
from scipy.interpolate import interp1d
from xlrd.sheet import ctype_text 
from scipy.signal import savgol_filter
from keras.models import Sequential, Model
from keras.layers import Dense, Reshape, BatchNormalization, MaxPool2D, Convolution2D, InputLayer, Flatten, Concatenate
from keras.utils import plot_model

pytesseract.pytesseract.tesseract_cmd = (os.path.join(os.getcwd(), "Tesseract-OCR\\tesseract.exe"))

curr_dir = os.getcwd()  
path_samples = os.path.join(curr_dir, "Data")   
ocr_data_file = os.path.join(path_samples, "OCR_dict.json")    
path_Post_Process = os.path.join(path_samples, "Post_Process") 
path_ref_images = os.path.join(path_Post_Process, "Reference_images") 
path_Interfaces = os.path.join(path_samples, "Interfaces") 
path_Threshold = os.path.join(path_samples, "Threshold") 
name_samples = next(os.walk(path_samples))[1] 



df_files = pd.DataFrame({'Sample': str,
                'Image_Name': [],
                'Image_ID':str,         
                'Image_Type': [],
                'Magnification': [],
                'Image_Location': [],
                'Image_Location_PostPocess': [],
                'Image_Location_Interface': []})

df_slice = pd.DataFrame({'Sample': str,
                'Image_ID': str,
                'Slice_ID':int,
                'Slice_Array': [],
                'Interface_Array': [],
                'Slice_Array_Threshold': []})

df_slice_porosity = pd.DataFrame({'Sample': str,
                'Image_ID': str,
                'Slice_ID':int,
                'Points_area': [],
                'Threads_area': [],
                'Clusters_area': [],
                'Total_area': []})

Hg_porosity_df = pd.DataFrame({'Sample': str,
                'Image_ID': str,
                'Slice_ID':int,
                'Radii': [],
                'Ps': [],
                'log_Radii': [],
                'log_Radii_new': [],
                'Ps_new': []})

def directory_check():
    flag = True
    ref_img_500 = os.path.join(path_ref_images,"ref_img_500.tif")
    ref_img_2000 = os.path.join(path_ref_images,"ref_img_2000.tif")
    
    if not os.path.isdir(path_samples):
        os.makedirs(path_samples)

    if not os.path.isdir(path_Post_Process):
        os.makedirs(path_Post_Process)
        
    if not os.path.isdir(path_Threshold):
        os.makedirs(path_Threshold)

    if not os.path.isdir(path_ref_images):
        os.makedirs(path_ref_images)
        flag = False
        
    if not os.path.isfile(ref_img_500) or not os.path.isfile(ref_img_2000):
        flag = False

    if not os.path.isdir(path_Interfaces):
        os.makedirs(path_Interfaces)

    if flag is True:
        print("All Directories Initialized Succesfully")
        
    else:
        print("Place the refenrence images for 500X (ref_img_500.tif) and 2000X (ref_img_2000.tif) pre-processing in folcer: ",path_ref_images)

def check_empty_json(standard_json):
    if not os.path.isfile(standard_json):
        print(standard_json + " does not exist")
        return(0)
    else:
        print("previous copy of ", standard_json, " exists")
        return(1)
        
def view1_image(imageA):
    fig = plt.figure(figsize=(18, 16), dpi= 80, facecolor='w', edgecolor='k')
    ax = fig.add_subplot(1, 2, 1)
    plt.imshow(imageA, cmap = plt.cm.gray)
    plt.axis("off")
    plt.show()
    
def view2_image(imageA, imageB):
    fig = plt.figure(figsize=(18, 16), dpi= 80, facecolor='w', edgecolor='k')
    ax = fig.add_subplot(1, 2, 1)
    plt.imshow(imageA, cmap = plt.cm.gray)
    plt.axis("off")
    ax = fig.add_subplot(1, 2, 2)
    plt.imshow(imageB, cmap = plt.cm.gray)
    plt.axis("off")
    plt.show()
     
        
def ocr(filename):
    img = cv2.imread(filename, cv2.IMREAD_UNCHANGED)
    height = img.shape[0]
    width = img.shape[1]
    microscope = "Other"
    OCR_Check_Complete = False
    
    ocr_phenom = img[int(0.95 * img.shape[0]):int(0.98 * img.shape[0]), int(0.68 * img.shape[1]):int(0.80 * img.shape[1])]
    ocr_lowres = img[int(0.93 * img.shape[0]):int(0.96 * img.shape[0]), int(0.72 * img.shape[1]):int(0.80 * img.shape[1])]
    
    list_500 = []
    list_2000 = []
    list_500_phenom = [str(i) for i in range(450, 550)] 
    list_2000_phenom = [str(i) for i in range(1950, 2050)] 
    list_500_lowres = [str(i) for i in range(18, 22)]
    list_2000_lowres = [str(i) for i in range(8, 12)]
    
    if height > 2000:
        ocr = ocr_phenom
        list_500 = list_500_phenom
        list_2000 = list_2000_phenom
        microscope = "Phenom"
        
    else:
        ocr = ocr_lowres
        list_500 = list_500_lowres
        list_2000 = list_2000_lowres
        microscope = "LowRes"

    ocr_text = pytesseract.image_to_string(ocr, config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789')
    
    if not ocr_text:
        ocr_text = "NONE"
    else: 
        ocr_text = str((re.findall('\d+', ocr_text))[0])
    
    if ocr_text in list_500:
        ocr_text = 500
        OCR_Check_Complete = True
        print("OCR Check complete = ", ocr_text)
        
    if ocr_text in list_2000:
        ocr_text = 2000   
        OCR_Check_Complete = True
        print("OCR Check complete = ", ocr_text)  
        
    if OCR_Check_Complete is False:
        print("OCR ERROR, Detected text = ", ocr_text)
        print("image location: ", filename)
        view1_image(ocr)
        ocr_text = input("***Enter the magnification manually : ")
        
    return ocr_text
        
        
        
        # def ocr(filename):
#     img = cv2.imread(filename, cv2.IMREAD_UNCHANGED)
#     ocr = img[int(0.95 * img.shape[0]):int(0.98 * img.shape[0]), int(0.68 * img.shape[1]):int(0.80 * img.shape[1])]
#     ocr_text = pytesseract.image_to_string(ocr, config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789')
#     if "500" not in ocr_text and "2000" not in ocr_text:
#         if not ocr_text:
#             ocr_text = "NONE"
#         print("OCR ERROR, Detected text = ", ocr_text)
#         print("image location: ", filename)
#         view1_image(ocr)
#         ocr_text = input("***Enter the magnification manually : ")

#     if ocr_text:
#         ocr_text = int((re.findall('\d+', ocr_text))[0])
#         print("OCR Check complete = ", ocr_text)
        
#     return ocr_text    

def df_files_update(df_files, start_ID = "JSC100", status = -1):
    OCR_dict = {} 
    OCR_dict_append = {}
    ID_No = 0  

    for root, dirs, files in os.walk(path_samples):
        for name in files:
            if os.path.basename(root) != "Reference_images":
                filename = os.path.join(root, name)
                if os.path.splitext(name)[1] == ".tif":
                    possible_sample_name_1 = os.path.basename(os.path.dirname(os.path.dirname(os.path.dirname(filename))))
                    possible_sample_name_2 = os.path.basename(os.path.dirname(os.path.dirname(filename)))

                    if possible_sample_name_1 in name_samples or possible_sample_name_2 in name_samples:                    
                        image_name = os.path.basename(filename)
                        image_location = filename
                        if possible_sample_name_1 in name_samples:
                            possible_sample_name = possible_sample_name_1
                            image_type_1 = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                            image_type_2 = os.path.basename(os.path.dirname(filename))
                            image_type = image_type_1 + "-" + image_type_2

                        else:
                            possible_sample_name = possible_sample_name_2
                            image_type = os.path.basename(os.path.dirname(filename))

                        ID_No = ID_No + 1
                        Image_ID = "JSC" + str(100 + ID_No)

                        if status == -1:
                            status = check_empty_json(ocr_data_file)

                        if status == 0:
                            key = image_name
                            magnification = ocr(image_location)
                            OCR_dict[key] = magnification

                        if status == 1:
                            with open(ocr_data_file) as json_file:
                                OCR_dict_append = json.load(json_file)
                                key = image_name
                                if key in OCR_dict_append:
                                    magnification = OCR_dict_append[key]
                                else:
                                    key = image_name
                                    magnification = ocr(image_location)
                                    OCR_dict_append[key] = magnification


                        df_files = df_files.append({'Sample': possible_sample_name, 
                                                    'Image_Name': image_name, 
                                                    'Image_ID': Image_ID,
                                                    'Image_Type': image_type,
                                                    'Magnification': magnification,
                                                    'Image_Location': image_location},  ignore_index=True)    

    if status == 0:
        print("(Status-0) OCR Detection complete..")
        json_write = json.dumps(OCR_dict)
        with open(ocr_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close()

    if status == 1:
        print("(Status-1) OCR Detection complete..")
        json_write = json.dumps(OCR_dict_append)
        with open(ocr_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close() 
            
    return df_files


def pre_process(ref_img_location, img, ocr_text, name):    
    
    ref_img = cv2.imread(ref_img_location, cv2.IMREAD_UNCHANGED)
    if ocr_text == 500:      
        rows = img.shape[0]
        cols = img.shape[1]
        strip_top = img[0:10, 0:int(cols)]
        avg_px_top = np.mean(strip_top, axis=0, dtype=np.int)
        
        for j in range(0, cols):  
            if avg_px_top[j] > 30:
                if j - 1 < 0:
                    for k in range(0, len(avg_px_top)):
                        avg_px_top[k] = 255     
                else:
                    avg_px_top[j] = avg_px_top[j-1]       
            
        for i in range(0, rows):
            for j in range(0, cols):                  
                if img[i, j] < avg_px_top[j] and avg_px_top[j] != 255: 
                    img[i, j] = 0
                    
                if img[i, j] > avg_px_top[j] and avg_px_top[j] != 255:
                    img[i, j] = img[i, j] - avg_px_top[j]

    ref_img_stat = int(ImageStat.Stat(Image.fromarray(ref_img)).mean[0])
    img_trial = Image.fromarray(img)
    ref_trial_stat = int(ImageStat.Stat(img_trial).mean[0])
    enhanced_img_stat = ref_trial_stat 

    while enhanced_img_stat is not ref_img_stat:
        if enhanced_img_stat < ref_img_stat:  
            enhancer = ImageEnhance.Brightness(img_trial)
            enhanced_img = enhancer.enhance(1.1)
            enhanced_img_stat = ImageStat.Stat(enhanced_img)
            enhanced_img_stat = int(enhanced_img_stat.mean[0])
            img_trial = enhanced_img

        if enhanced_img_stat > ref_img_stat:  
            enhancer = ImageEnhance.Brightness(img_trial)
            enhanced_img = enhancer.enhance(0.99)
            enhanced_img_stat = ImageStat.Stat(enhanced_img)
            enhanced_img_stat = int(enhanced_img_stat.mean[0])
            img_trial = enhanced_img

        if abs(enhanced_img_stat - ref_img_stat) < 3:  
            img_trial.save(os.path.join(path_Post_Process, name))
            return ref_img

    img_trial.save(os.path.join(path_Post_Process, name))
    
    
def df_files_preprocess(df_files, status = -1):
    ref_img_500 = os.path.join(path_ref_images,"ref_img_500.tif")
    ref_img_2000 = os.path.join(path_ref_images,"ref_img_2000.tif")  
    PreProcess_data_file = os.path.join(path_samples, "PreProcess_dict.json")
    
    PreProcess_dict = {}
    PreProcess_dict_append = {}
    Image_Location_PostPocess_col = []

    for index, row in df_files.iterrows():
        Image_ID = row['Image_ID']
        Image_Name = row['Image_Name']
        Image_Location = row['Image_Location']
        Magnification = row['Magnification']
        img = cv2.imread(Image_Location, cv2.IMREAD_UNCHANGED)       
    
        if status == -1:
            status = check_empty_json(PreProcess_data_file)

        if status == 0:
            print("(Status-New) Processing", Image_ID, "with Magnification", Magnification," ..")  
            if int(Magnification) == 500:
                pre_process(ref_img_500, img, Magnification, Image_ID + ".tif")
                Image_Location_PostPocess = os.path.join(path_Post_Process, Image_ID + ".tif")
                Image_Location_PostPocess_col.append(Image_Location_PostPocess)

            if int(Magnification) == 2000:
                pre_process(ref_img_2000, img, Magnification, Image_ID + ".tif")
                Image_Location_PostPocess = os.path.join(path_Post_Process, Image_ID + ".tif")
                Image_Location_PostPocess_col.append(Image_Location_PostPocess)

            PreProcess_dict[Image_Name] = Image_Location_PostPocess  
        
        if status == 1:
            print("(Status-Old) Processing", Image_ID, "with Magnification", Magnification," ..")  
            with open(PreProcess_data_file) as json_file:
                PreProcess_dict_append = json.load(json_file)
                row, col = np.where(df_files == Image_Name)
                path = os.path.join(path_Post_Process, df_files.iloc[row[0], 2] + ".tif")

                if Image_Name in PreProcess_dict_append and os.path.isfile(path):
                    Image_Location_PostPocess = PreProcess_dict_append[Image_Name]
                    if Image_Location_PostPocess != path:
                        Image_Location_PostPocess = path
                    Image_Location_PostPocess_col.append(Image_Location_PostPocess)

                else:
                    if int(Magnification) == 500:
                        pre_process(ref_img_500, img, Magnification, Image_ID + ".tif")
                        Image_Location_PostPocess = os.path.join(path_Post_Process, Image_ID + ".tif")
                        Image_Location_PostPocess_col.append(Image_Location_PostPocess)

                    if int(Magnification) == 2000:
                        pre_process(ref_img_2000, img, Magnification, Image_ID + ".tif")
                        Image_Location_PostPocess = os.path.join(path_Post_Process, Image_ID + ".tif")
                        Image_Location_PostPocess_col.append(Image_Location_PostPocess)

                    PreProcess_dict_append[Image_Name] = Image_Location_PostPocess
        
    df_files['Image_Location_PostPocess'] = Image_Location_PostPocess_col

    if status == 0:
        print("(Status-0) PreProcessing complete..")
        json_write = json.dumps(PreProcess_dict)
        with open(PreProcess_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close()


    if status == 1:
        print("(Status-1) PreProcessing complete..")
        json_write = json.dumps(PreProcess_dict_append)
        with open(PreProcess_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close() 
            
    return df_files
            
def interface_detection(roi):
    rows = roi.shape[0]
    cols = roi.shape[1]
    blur = cv2.GaussianBlur(roi, (5, 5), 0)
    ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    interface = cv2.Canny(th, 100, 200)
    inter = np.zeros([rows, cols], dtype=int)
    
    for i in range(0, rows):
        for j in range(0, cols):
            if interface[i, j] == 255:
                inter = interface[0:i + 1, j:j + 1].ravel()
                inter_mean = np.mean(inter[0:i])

                if inter_mean < 0.1:
                    interface[i, j] = 250
  
    interface[interface != 250] = 0
    interface[interface == 250] = 255
    return interface


def df_files_interface(df_files, status = -1):
    Interface_data_file = os.path.join(path_samples, "Interface_dict.json")
    Interface_dict = {}
    Interface_dict_append = []
    Image_Location_Interface_col = []

    for index, row in df_files.iterrows():
        Image_ID = row['Image_ID']
        Image_Location_PostPocess  = row['Image_Location_PostPocess']
        Magnification = row['Magnification']
        Image_Name = row['Image_Name']

        if status == -1:
            status = check_empty_json(Interface_data_file)

        if status == 0:
            if Magnification == 500:
                img = cv2.imread(Image_Location_PostPocess, cv2.IMREAD_UNCHANGED)
                print("(Status-New) Generating Top Interface File for 500X image: ", Image_ID, "...")
                interface = interface_detection(img[0:int(0.30 * img.shape[0]), 0:int(img.shape[1])])
                Image_Location_Interface = os.path.join(path_Interfaces, Image_ID + ".tif")
                Image_Location_Interface_col.append(Image_Location_Interface)
                cv2.imwrite(Image_Location_Interface, interface)

            if Magnification == 2000:
                Image_Location_Interface = "<No_Interface>"
                print("(Status-New) No Interface File generated for 2000X image: ", Image_ID, "...")
                Image_Location_Interface_col.append(Image_Location_Interface)

            Interface_dict[Image_Name] = Image_Location_Interface  

        if status == 1:
            with open(Interface_data_file) as json_file:
                Interface_dict_append = json.load(json_file)
                row, col = np.where(df_files == Image_Name)

                if Image_Name in Interface_dict_append:
                    Image_Location_Interface = Interface_dict_append[Image_Name]
                    if not os.path.isfile(Image_Location_Interface):
                        Image_Location_Interface = os.path.join(path_Interfaces, Image_ID + ".tif")
                    Image_Location_Interface_col.append(Image_Location_Interface)

                else:
                    if Magnification == 500:
                        img = cv2.imread(Image_Location_PostPocess, cv2.IMREAD_UNCHANGED)
                        interface = interface_detection(img[0:int(0.30 * img.shape[0]), 0:int(img.shape[1])])
                        Image_Location_Interface = os.path.join(path_Interfaces, Image_ID + ".tif")
                        Image_Location_Interface_col.append(Image_Location_Interface)
                        cv2.imwrite(Image_Location_Interface, interface)

                    if Magnification == 2000:
                        Image_Location_Interface = "<No_Interface>"
                        Image_Location_Interface_col.append(Image_Location_Interface)

                    Interface_dict_append[Image_Name] = Image_Location_Interface


    df_files['Image_Location_Interface'] = Image_Location_Interface_col

    if status == 0:
        print("(Status-0) Interface Extraction complete..")
        json_write = json.dumps(Interface_dict)
        with open(Interface_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close()

    if status == 1:
        print("(Status-1) Interface Extraction complete..")
        json_write = json.dumps(Interface_dict_append)
        with open(Interface_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close() 
 
    return df_files

def df_files_threshold(df_files, status = -1):
    # 1st Operation
    def image_dilate(img):
        img_white_matrix = np.copy(img)
        blur = cv2.GaussianBlur(img_white_matrix, (5, 5), 0)
        ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        morph = th.copy()
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
        morph = cv2.morphologyEx(morph, cv2.MORPH_CLOSE, kernel)
        morph = cv2.morphologyEx(morph, cv2.MORPH_OPEN, kernel)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        erosion = cv2.erode(img_white_matrix, kernel, iterations = 1)
        dilate = cv2.dilate(erosion, kernel, iterations = 2)
        erosion = cv2.erode(dilate, kernel, iterations = 1)
        img_white_matrix_dilate = cv2.dilate(erosion, kernel, iterations = 1)
        return img_white_matrix_dilate
    
    

    # 2nd Operation
    def delete_non_matrix(img_original):
        img = np.copy(img_original)
        neighbours = np.zeros((4,4))
        rows = img.shape[0]
        cols = img.shape[1]
        inter = []
        inter_mean = 0
        for j in range(0, cols):
            for i in range(0, rows):
                if img[i, j] == 255:
                    inter = img[0:i + 1, j:j + 1].ravel()
                    inter_mean = np.mean(inter[0:i])       
                    if inter_mean < 2.0:
                        img[0:i + 1, j:j + 1] = 120
                    break

        for j in range(0, cols):
            for i in range(rows-1, 0, -1):
                if img[i, j] == 255:
                    inter = img[i + 1:rows, j:j + 1].ravel()
                    inter_mean = np.mean(inter[0:i])    
                    if inter_mean < 2.0 and not math.isnan(inter_mean):
                        img[i + 1:rows, j:j + 1] = 120
                    break  

        return img

    # 3.1th Operation
    def find_matrix_threshhold(img_part):
        naux = 0
        naux_color = 0 
        threshold = 0
        threshold_offset = 20
        threshold_fine = []
        ascending_pixel = []
        curr_max_pixel = 0
        curr_max_pixel_index = 0
        pix_width = img_part.shape[0]
        pix_height = img_part.shape[1]

        for i in range(256):
            threshold_fine.append(0)
            ascending_pixel.append(0)

        for i in range(pix_width):
            for j in range(pix_height):
                pix = int(img_part[i,j])
                threshold_fine[pix] = threshold_fine[pix] + 1

            if naux <= threshold_fine[pix]: 
                naux = threshold_fine[pix] 
                naux_color = pix

        for i in range(len(threshold_fine)):
            curr_max_pixel = np.max(threshold_fine)
            curr_max_pixel_index_tuple = np.where(threshold_fine == curr_max_pixel)
            curr_max_pixel_index = curr_max_pixel_index_tuple[0][0]  
            ascending_pixel[i] = curr_max_pixel_index
            threshold_fine[curr_max_pixel_index] = 0

        for i in range(len(ascending_pixel)):
            if ascending_pixel[i] > 150:
                threshold = ascending_pixel[i] - threshold_offset
                break
            else:
                threshold = 150      

        return threshold

    # 3rd Operation
    def matrix_find_thresh(img):
        img_copy = np.copy(img)
        height_image = img_copy.shape[0]
        width_image = img_copy.shape[1]
        div = 5     

        height_part = int(height_image/div)
        width_part = int(width_image/div)

        dif_height = height_image - int(height_part*div)
        dif_width = width_image - int(width_part*div)

        i = 0
        while i<height_image:
            j = 0
            while j<width_image:
                if (i + height_part <= height_image) and (j + width_part <= width_image):

                    img_part = img_copy[i:i+height_part, j:j+width_part]
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[i:i+height_part, j:j+width_part] = th

                if j + width_part > width_image:
                    img_part = img_copy[0:height_part, j:j+dif_width] 
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[0:height_part, j:j+dif_width]  = th

                if i + height_part > height_image:
                    img_part = img_copy[i:i+dif_height, 0:dif_width] 
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[i:i+dif_height, 0:dif_width] = th

                j = j + width_part
            i = i + height_part
            
        view1_image(img_copy)

        return img_copy

    
    Threshold_data_file = os.path.join(path_samples, "Threshold_dict.json")
    Threshold_dict = {}
    Threshold_dict_append = {}
    Image_Location_Threshold_col = []
    
    for index, row in df_files.iterrows():
        Image_ID = row['Image_ID']
        Image_Name = row['Image_Name']
        Image_Location_PostPocess = row['Image_Location_PostPocess']
        img = cv2.imread(Image_Location_PostPocess, cv2.IMREAD_UNCHANGED) 
        
        if status == -1:
            status = check_empty_json(Threshold_data_file)
            
        if status == 0:
            print("(Status-New) Thresholding", Image_ID)          
            img_white_matrix_dilate = image_dilate(img)
            img_white_matrix_dilate_without_matrix = delete_non_matrix(img_white_matrix_dilate)
            img_after_find_matrix = matrix_find_thresh(img_white_matrix_dilate_without_matrix) 
            
            Image_Location_Threshold = os.path.join(path_Threshold, Image_ID + ".tif")
            save_img_after_find_matrix = Image.fromarray(img_after_find_matrix)
            save_img_after_find_matrix.save(Image_Location_Threshold)
            
            Image_Location_Threshold_col.append(Image_Location_Threshold)      
            Threshold_dict[Image_Name] = Image_Location_Threshold  
            
            
        if status == 1: 
            with open(Threshold_data_file) as json_file:
                Threshold_dict_append = json.load(json_file)
                row, col = np.where(df_files == Image_Name)
                path = os.path.join(path_Threshold, df_files.iloc[row[0], 2] + ".tif")

                if Image_Name in Threshold_dict_append and os.path.isfile(path):
                    Image_Location_Threshold = Threshold_dict_append[Image_Name]
                    if Image_Location_Threshold != path:
                        Image_Location_Threshold = path
                    Image_Location_Threshold_col.append(Image_Location_Threshold)

                else:
                    print("(Status-Append) Thresholding", Image_ID)          
                    img_white_matrix_dilate = image_dilate(img)
                    img_white_matrix_dilate_without_matrix = delete_non_matrix(img_white_matrix_dilate)
                    img_after_find_matrix = matrix_find_thresh(img_white_matrix_dilate_without_matrix)   
                    
                    Image_Location_Threshold = os.path.join(path_Threshold, Image_ID + ".tif")
                    save_img_after_find_matrix = Image.fromarray(img_after_find_matrix)
                    save_img_after_find_matrix.save(Image_Location_Threshold)
                    
                    Image_Location_Threshold_col.append(Image_Location_Threshold)      
                    Threshold_dict_append[Image_Name] = Image_Location_Threshold
    
    
    df_files['Image_Location_Threshold'] = Image_Location_Threshold_col

    if status == 0:
        print("(Status-0) Thresholding complete..")
        json_write = json.dumps(Threshold_dict)
        with open(Threshold_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close()


    if status == 1:
        print("(Status-1) Thresholding complete..")
        json_write = json.dumps(Threshold_dict_append)
        with open(Threshold_data_file, 'w') as json_file:
            json_file.write(json_write)
            json_file.close()




def df_files_check(df_files, missing = False):
    for index, row in df_files.iterrows():
        Image_ID = row['Image_ID']
        Image_Location_PostPocess  = row['Image_Location_PostPocess']
        Image_Location_Interface = row['Image_Location_Interface']
        Image_Location_Threshold = row['Image_Location_Threshold']
        Magnification = row['Magnification']
        Image_Name = row['Image_Name']

        if not os.path.isfile(Image_Location_PostPocess):
            missing = True
            print("\nThe following file did not undergo pre-process")
            print("Path: ", df_files.Image_Location[index])
            print("ID: ", df_files.Image_Unique_ID[index])
            print("Magnification: ", df_files.Magnification[index])
            print("Post Proces Location: ", df_files.Image_Location_PostPocess[index])
            
            
        if not os.path.isfile(Image_Location_Interface) and int(Magnification) != 2000:
            missing = True
            print("\nThe following file do not have interface: ", Image_Location_Interface)
            print("Path: ", df_files.Image_Location[index])
            print("ID: ", df_files.Image_ID[index])
            print("Magnification: ", df_files.Magnification[index])
            print("Post Proces Location: ", df_files.Image_Location_PostPocess[index])
            
            
        if not os.path.isfile(Image_Location_Threshold):
            missing = True
            print("\nThe following file do not have threshold: ", Image_Location_Threshold)
            print("Path: ", df_files.Image_Location[index])
            print("ID: ", df_files.Image_ID[index])
            print("Magnification: ", df_files.Magnification[index])
            print("Post Proces Location: ", df_files.Image_Location_Threshold[index])
        
    if missing is False:
        print("All files were processed..")



def delete_non_matrix(img_original):
    img = np.copy(img_original)
    neighbours = np.zeros((4,4))
    rows = img.shape[0]
    cols = img.shape[1]
    inter = []
    inter_mean = 0
    for j in range(0, cols):
        for i in range(0, rows):
            if img[i, j] == 255:
                inter = img[0:i + 1, j:j + 1].ravel()
                inter_mean = np.mean(inter[0:i])       
                if inter_mean < 2.0:
                    img[0:i + 1, j:j + 1] = 120
                break
    
    for j in range(0, cols):
        for i in range(rows-1, 0, -1):
            if img[i, j] == 255:
                inter = img[i + 1:rows, j:j + 1].ravel()
                inter_mean = np.mean(inter[0:i])    
                if inter_mean < 2.0 and not math.isnan(inter_mean):
                    img[i + 1:rows, j:j + 1] = 120
                break         
    return img

def image_dilate(img):
    img_white_matrix = np.copy(img)
    blur = cv2.GaussianBlur(img_white_matrix, (5, 5), 0)
    ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    morph = th.copy()
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
    morph = cv2.morphologyEx(morph, cv2.MORPH_CLOSE, kernel)
    morph = cv2.morphologyEx(morph, cv2.MORPH_OPEN, kernel)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    erosion = cv2.erode(img_white_matrix, kernel, iterations = 1)
    dilate = cv2.dilate(erosion, kernel, iterations = 2)
    erosion = cv2.erode(dilate, kernel, iterations = 1)
    img_white_matrix_dilate = cv2.dilate(erosion, kernel, iterations = 1) 
    
    return img_white_matrix_dilate

def find_matrix_threshhold(img_part):
    naux = 0
    naux_color = 0 
    threshold = 0
    threshold_offset = 20
    threshold_fine = []
    ascending_pixel = []
    curr_max_pixel = 0
    curr_max_pixel_index = 0
    pix_width = img_part.shape[0]
    pix_height = img_part.shape[1]
    
    for i in range(256):
        threshold_fine.append(0)
        ascending_pixel.append(0)
        
    for i in range(pix_width):
        for j in range(pix_height):
            pix = int(img_part[i,j])
            threshold_fine[pix] = threshold_fine[pix] + 1

        if naux <= threshold_fine[pix]: 
            naux = threshold_fine[pix] 
            naux_color = pix

    for i in range(len(threshold_fine)):
        curr_max_pixel = np.max(threshold_fine)
        curr_max_pixel_index_tuple = np.where(threshold_fine == curr_max_pixel)
        curr_max_pixel_index = curr_max_pixel_index_tuple[0][0]  
        ascending_pixel[i] = curr_max_pixel_index
        threshold_fine[curr_max_pixel_index] = 0
    
    for i in range(len(ascending_pixel)):
        if ascending_pixel[i] > 150:
            threshold = ascending_pixel[i] - threshold_offset
            break
        else:
            threshold = 150      
       
    return threshold
    

def matrix_find_thresh(img):
    img_copy = np.copy(img)
    height_image = img_copy.shape[0]
    width_image = img_copy.shape[1]
    div = 5     
    
    height_part = int(height_image/div)
    width_part = int(width_image/div)

    dif_height = height_image - int(height_part*div)
    dif_width = width_image - int(width_part*div)
    
    i = 0
    while i<height_image:
        j = 0
        while j<width_image:
            if (i + height_part <= height_image) and (j + width_part <= width_image):
                
                img_part = img_copy[i:i+height_part, j:j+width_part]
                thresh_img_part = find_matrix_threshhold(img_part)
                ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                img_copy[i:i+height_part, j:j+width_part] = th

            if j + width_part > width_image:
                img_part = img_copy[0:height_part, j:j+dif_width] 
                thresh_img_part = find_matrix_threshhold(img_part)
                ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                img_copy[0:height_part, j:j+dif_width]  = th
                
            if i + height_part > height_image:
                img_part = img_copy[i:i+dif_height, 0:dif_width] 
                thresh_img_part = find_matrix_threshhold(img_part)
                ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                img_copy[i:i+dif_height, 0:dif_width] = th
            
            j = j + width_part
        i = i + height_part
            
    return img_copy


def get_interface_boundaries(img):
    horizontal_kernel = np.ones((1,2048), np.uint8)  # note this is a horizontal kernel
    kernel = np.ones((1,20), np.uint8)
    top_row = 0
    bottom_row = 0
    
    top_section = img[0:int(0.30 * img.shape[0]), 0:int(img.shape[1])]    
    rows = top_section.shape[0]
    cols = top_section.shape[1]
    blur = cv2.GaussianBlur(top_section, (5, 5), 0)
    ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    top_section_interface = cv2.Canny(th, 100, 200)
    inter = np.zeros([rows, cols], dtype=int)
    for i in range(0, rows):
        for j in range(0, cols):
            if top_section_interface[i, j] == 255:
                inter = top_section_interface[0:i + 1, j:j + 1].ravel()
                inter_mean = np.mean(inter[0:i])

                if inter_mean < 0.1:
                    top_section_interface[i, j] = 250
                    
    top_section_interface[top_section_interface != 250] = 0
    top_section_interface[top_section_interface == 250] = 255
    
    dil_top_section_interface = cv2.dilate(top_section_interface, kernel, iterations=1)
    ero_top_section_interface = cv2.erode(dil_top_section_interface, kernel, iterations=1) 
    
    count = np.zeros(rows, dtype=int)
    for i in range(0, rows):
        for j in range(0, cols):
            if top_section_interface[i, j] == 255:
                count[i] = count[i] + 1
    
    max_count_row = np.argmax(count)
    for i in range(max_count_row, rows):
        if count[i] == 0:
            top_row = i
            break

    bottom_section = img[int(0.70 * img.shape[0]):img.shape[0] - 135, 0:int(img.shape[1])]    
    rows = bottom_section.shape[0]
    cols = bottom_section.shape[1]
    
    ret, th1 = cv2.threshold(bottom_section, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    inter = th1[rows - 1:rows, 0:cols].ravel() 
    count = np.sum(inter)/len(inter)
    if count > 200:
        bottom_row = img.shape[0] - 135
               
    else: 
        blur = cv2.GaussianBlur(bottom_section, (5, 5), 0)
        ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        bottom_section_interface = cv2.Canny(th, 100, 200)
        inter = np.zeros([rows, cols], dtype=int)
        for i in range(0, rows):
            for j in range(0, cols):
                if bottom_section_interface[i, j] == 255:
                    inter = bottom_section_interface[i:rows, j:j + 1].ravel()           
                    inter_count = np.count_nonzero(inter[0:rows-i])
                    if inter_count <= 1:
                        bottom_section_interface[i, j] = 250

        bottom_section_interface[bottom_section_interface != 250] = 0
        bottom_section_interface[bottom_section_interface == 250] = 255   

        count = np.zeros(rows, dtype=int)
        for i in range(0, rows):
            for j in range(0, cols):    
                if bottom_section_interface[i, j] == 255:
                    count[i] = count[i] + 1   

        max_count_row = np.argmax(count)
        for i in range(200, max_count_row):
            if count[i] > 5:
                bottom_row = i
                break

        bottom_row = rows - bottom_row
        bottom_row = img.shape[0] - bottom_row - 135
        
    return top_row, bottom_row

def slice_image_interface(filename, x1, x2, diam):
    name = os.path.basename(filename)
    interface_location = os.path.join(path_Interfaces, name)   
    interface_img = cv2.imread(interface_location, cv2.IMREAD_UNCHANGED)       
    slice_interface = interface_img[0:interface_img.shape[0], x1:x2] 
    
    sum_i = 0
    n = 0
    
    for i in range(0, interface_img.shape[0]):
        for j in range(0, x2-x1):    
            if slice_interface[i,j] == 255:
                sum_i = sum_i + i
                n = n + 1
            
    avg_i = int(sum_i/n)
    start_row = avg_i - int(diam/2)
    end_row = avg_i + int(diam/2)
    
    if start_row < 0:
        end_row = end_row + abs(start_row)
        start_row = 0
      
    slice_interface = slice_interface[start_row : end_row, 0 : diam]
    return(slice_interface)
    
    
def slice_image(filename, top, bottom, diam, no_of_slices, overlap):  
    img = cv2.imread(filename, cv2.IMREAD_UNCHANGED) 
    img_white_matrix_dilate = image_dilate(img)
    img_white_matrix_dilate_without_matrix = delete_non_matrix(img_white_matrix_dilate)
    img_after_find_matrix = matrix_find_thresh(img_white_matrix_dilate_without_matrix)
    slices_main = img[top:bottom, 0:int(img.shape[1])] 
    slices_main_thresh = img_after_find_matrix[top:bottom, 0:int(img.shape[1])] 
    
    slice_centres_x = 0
    slice_centres_y = 0
    slice_sq_dim = diam  
    slices = [np.zeros((slice_sq_dim, slice_sq_dim)) for x in range(no_of_slices)]
    slices_thresh = [np.zeros((slice_sq_dim, slice_sq_dim)) for x in range(no_of_slices)]
    slices_interface = [np.zeros((slice_sq_dim, slice_sq_dim)) for x in range(no_of_slices)]
    
    blur = cv2.GaussianBlur(img, (5, 5), 0)
    ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    img = cv2.Canny(th, 100, 200)
    
    if overlap == "No":
        no_of_slices = 12
        rows = bottom - top
        cols = img.shape[1]
        row_increment = int(rows / 3)
        col_increment = int(cols / 4)
        sl_no = -1

        for i in range(top, bottom, row_increment):
            for j in range(0, cols, col_increment):  
                if i + row_increment > bottom:
                    break                

                sl_no = sl_no + 1
                index_i = int((i - top) / row_increment)
                index_j = int(j / col_increment)

                rand_row_low = i + int(slice_sq_dim/2)
                rand_row_high = i + row_increment - int(slice_sq_dim/2)
                rand_col_low = j + int(slice_sq_dim/2)
                rand_col_high = j + col_increment - int(slice_sq_dim/2)

                if rand_row_low > rand_row_high:
                    diff = rand_row_low - rand_row_high
                    rand_row_low = rand_row_low - int(diff/2) - np.random.randint(1, 5, size=1)[0]
                    rand_row_high = rand_row_high + int(diff/2) + np.random.randint(1, 5, size=1)[0] 

                if rand_col_low > rand_col_high:
                    diff = rand_col_low - rand_col_high
                    rand_col_low = rand_col_low - int(diff/2) - np.random.randint(1, 5, size=1)[0]
                    rand_col_high = rand_col_high + int(diff/2) + np.random.randint(1, 5, size=1)[0] 

                rand_row = (np.random.randint(rand_row_low, rand_row_high, size=1))[0]
                rand_col = (np.random.randint(rand_col_low, rand_col_high, size=1))[0]
                slice_centres_x = rand_col
                slice_centres_y = rand_row

                cv2.circle(img, (slice_centres_x,slice_centres_y), 2, (0,255,0), -1)
                x1 = slice_centres_x - int( slice_sq_dim / 2 )
                y1 = slice_centres_y - int( slice_sq_dim / 2 ) 
                x2 = slice_centres_x + int( slice_sq_dim / 2 )
                y2 = slice_centres_y + int( slice_sq_dim / 2 )

                slices_interface[sl_no] =  slice_image_interface(filename, x1, x2, diam)       
                slices[sl_no] = slices_main[y1:y2, x1:x2]
                slices_thresh[sl_no] = slices_main_thresh[y1:y2, x1:x2]
    
    elif overlap == "Yes":
        
        for sl_no in range(no_of_slices):
            rand_row = (np.random.randint(top + slice_sq_dim, bottom - slice_sq_dim, size=1))[0]
            rand_col = (np.random.randint(slice_sq_dim / 2, img.shape[1], size=1))[0]
            slice_centres_x = rand_col
            slice_centres_y = rand_row
        
            cv2.circle(img, (slice_centres_x,slice_centres_y), 2, (0,255,0), -1)
            
            x1 = slice_centres_x - int( slice_sq_dim / 2 )
            y1 = slice_centres_y - int( slice_sq_dim / 2 ) 
            x2 = slice_centres_x + int( slice_sq_dim / 2 )
            y2 = slice_centres_y + int( slice_sq_dim / 2 )
            
            while y2 > slices_main.shape[0]:
                y1 = y1 - (np.random.randint(1, 100, size=1))[0]
                y2 = y1 + slice_sq_dim
                
            while x2 > slices_main.shape[1]:
                x1 = x1 - (np.random.randint(1, 100, size=1))[0]
                x2 = x1 + slice_sq_dim
            
            slices_interface[sl_no] =  slice_image_interface(filename, x1, x2, diam)       
            slices[sl_no] = slices_main[y1:y2, x1:x2]
            slices_thresh[sl_no] = slices_main_thresh[y1:y2, x1:x2]
  
    return slices, slices_thresh, slices_interface


def df_slice_update(df_files, df_slice, slice_sq_dim = 224, overlap = "Yes", no_of_slices = 20):
    df_slice = pd.DataFrame({'Sample': str,
               'Image_ID': str,
               'Slice_ID':int,
               'Slice_Array': [],
               'Interface_Array': [],
               'Slice_Array_Threshold': []})
    
    Slices_data_file = os.path.join(path_samples, "Slices_dataframe.json")
    indx_row = 0
    continue_operation = False
    
    if os.path.isfile(Slices_data_file): 
        print("Loading the latest slice data..(Delete 'Slices_dataframe.json' to start over)")
        
        df_slice = pd.read_json(Slices_data_file, orient = 'columns') 
        df_slice.sort_index(inplace=True)
        
        display(df_slice)
        
        required_number = df_files[df_files.Magnification == 500].shape[0] * no_of_slices
        current_number = df_slice.shape[0]
        
        if required_number > current_number:
            print("Completing the slicing operations for remaining images of exisitng file")
            continue_operation = True  
        else:
            return df_slice

    else:
        print("No previously saved slice data found, beginning NEW slicing operation..")
        for index, row in df_files.iterrows():
            Sample = row['Sample']
            Image_ID = row['Image_ID']
            magnification = row['Magnification']
            image_location_PostProcess = row["Image_Location_PostPocess"]
            
            Sample_col = []
            Image_ID_col = []
            Slice_ID_col = []
            Slice_Array_col = []
            Interface_Array_col = []
            Slice_Array_Threshold_col = []   
            
            df_slice_temp = pd.DataFrame({'Sample': str,
               'Image_ID': str,
               'Slice_ID':int,
               'Slice_Array': [],
               'Interface_Array': [],
               'Slice_Array_Threshold': []})

            if magnification == 500:
                img = cv2.imread(image_location_PostProcess, cv2.IMREAD_UNCHANGED)
                top, bottom = get_interface_boundaries(img)
                slices, slices_thresh, slices_interface = slice_image(image_location_PostProcess, top, bottom, slice_sq_dim, no_of_slices, overlap)
                for n, arr in enumerate(slices):
                    Sample_col.append(Sample)
                    Image_ID_col.append(Image_ID)
                    Slice_ID_col.append(n+1)
                    Slice_Array_col.append(arr)                    
                    Interface_Array_col.append(slices_interface[n])
                    Slice_Array_Threshold_col.append(slices_thresh[n])
    
                df_slice_temp['Sample'] = pd.Series(Sample_col)
                df_slice_temp['Image_ID'] = pd.Series(Image_ID_col)
                df_slice_temp['Slice_ID'] = pd.Series(Slice_ID_col)
                df_slice_temp['Slice_Array'] = pd.Series(Slice_Array_col)
                df_slice_temp['Interface_Array'] = pd.Series(Interface_Array_col)
                df_slice_temp['Slice_Array_Threshold'] = pd.Series(Slice_Array_Threshold_col)
                df_slice = df_slice.append(df_slice_temp, ignore_index = True)
                df_slice.to_json(Slices_data_file, orient='columns')
                
                print("Sicing -> ", Image_ID, " Complete and updated in the file")
                
        return df_slice
                
    if continue_operation is True:  
        for index, row in df_files.iterrows():          
            Sample = row['Sample']
            Image_ID = row['Image_ID']
            magnification = row['Magnification']
            image_location_PostProcess = row["Image_Location_PostPocess"]
            
            Sample_col = []
            Image_ID_col = []
            Slice_ID_col = []
            Slice_Array_col = []
            Interface_Array_col = []
            Slice_Array_Threshold_col = []
            
            df_slice_temp = pd.DataFrame({'Sample': str,
               'Image_ID': str,
               'Slice_ID':int,
               'Slice_Array': [],
               'Interface_Array': [],
               'Slice_Array_Threshold': []})

            pass_condition = False
            slice_id_list = df_slice['Image_ID'].tolist()
            if(Image_ID not in slice_id_list):
                pass_condition = True
            else:
                pass_condition = False
                
            if pass_condition:
                if magnification == 500:
                    img = cv2.imread(image_location_PostProcess, cv2.IMREAD_UNCHANGED)
                    top, bottom = get_interface_boundaries(img)
                    slices, slices_thresh, slices_interface = slice_image(image_location_PostProcess, top, bottom, slice_sq_dim, no_of_slices, overlap)
                    for n, arr in enumerate(slices):
                        Sample_col.append(Sample)
                        Image_ID_col.append(Image_ID)
                        Slice_ID_col.append(n+1)
                        Slice_Array_col.append(arr)
                        Interface_Array_col.append(slices_interface[n])
                        Slice_Array_Threshold_col.append(slices_thresh[n])

                    df_slice_temp['Sample'] = pd.Series(Sample_col)
                    df_slice_temp['Image_ID'] = pd.Series(Image_ID_col)
                    df_slice_temp['Slice_ID'] = pd.Series(Slice_ID_col)
                    df_slice_temp['Slice_Array'] = pd.Series(Slice_Array_col)
                    df_slice_temp['Interface_Array'] = pd.Series(Interface_Array_col)
                    df_slice_temp['Slice_Array_Threshold'] = pd.Series(Slice_Array_Threshold_col)
                    df_slice = df_slice.append(df_slice_temp, ignore_index = True)
                    df_slice.to_json(Slices_data_file, orient='columns')

                    print("Sicing -> ", Image_ID, " Complete and updated in the file")  
                    
        return df_slice
    
def df_slice_file_display():
    Slices_data_file = os.path.join(path_samples, "Slices_dataframe.json")
    if os.path.isfile(Slices_data_file): 
        print("Loading the latest (Temporary/Full) slice data..")
        df_slice_new = pd.read_json(Slices_data_file, orient = 'columns') 
        df_slice_new.sort_index(inplace=True)
        print("The slice file at: ", Slices_data_file, " has the following last rows: ")
        display(df_slice_new.tail(3)) 
    return df_slice_new


def df_slice_check(df_slice, slice_sq_dim = 224):
    missing = False
    if df_slice.empty:
        print("The slices dataframe is either empty or incomplete..")

    else:
        for index, row in df_slice.iterrows():
            Sample = row['Sample']
            Image_ID = row['Image_ID']
            Slice_ID = row['Slice_ID']
            Slice_Array = row["Slice_Array"]
            Interface_Array = row["Interface_Array"]
            Slice_Array_Threshold = row["Slice_Array_Threshold"]

            Slice_Array_new = np.asarray(Slice_Array, dtype='uint8')
            Interface_Array_new = np.asarray(Interface_Array, dtype='uint8')
            Slice_Array_Threshold_new = np.asarray(Slice_Array_Threshold, dtype='uint8')

            if Slice_Array_new.shape[0] == Slice_Array_new.shape[1] and Slice_Array_new.shape[0] == slice_sq_dim:
                pass
            else:
                missing = True
                print("Error in slice file of Image_ID: ", Image_ID, " and Slice_ID: ", Slice_ID)

            if Interface_Array_new.shape[0] == Interface_Array_new.shape[1] and Interface_Array_new.shape[0] == slice_sq_dim:
                pass
            else:
                missing = True
                print("Error in slice file of Image_ID: ", Image_ID, " and Slice_ID: ", Slice_ID)

            if Slice_Array_Threshold_new.shape[0] == Slice_Array_Threshold_new.shape[1] and Slice_Array_Threshold_new.shape[0] == slice_sq_dim:
                pass
            else:
                missing = True
                print("Error in slice file of Image_ID: ", Image_ID, " and Slice_ID: ", Slice_ID)
                
    if missing is False:
        print("All files were sliced correctly..")
        
        
        
        
def get_nearest_neighbours(img, pos_i, pos_j, pix_block_size):
    top_NN_i = []
    top_NN_j = []
    left_NN_i = []
    left_NN_j = []
    bottom_NN_i = []
    bottom_NN_j = []
    right_NN_i = []
    right_NN_j = []
    
    for i in range(pix_block_size+2):
        top_NN_i.append(pos_i - 1)
        left_NN_j.append(pos_j - 1)
        bottom_NN_i.append(pos_i + pix_block_size)
        right_NN_j.append(pos_j + pix_block_size)
         
    for i in range(-1,pix_block_size+1,1):
        top_NN_j.append(pos_j + i)
        left_NN_i.append(pos_i + i)
        bottom_NN_j.append(pos_j + i)
        right_NN_i.append(pos_i + i)
        
    top_NN = []
    left_NN = []
    bottom_NN = []
    right_NN = []
    
    for i in range(pix_block_size+2):
        top_NN.append((top_NN_i[i],top_NN_j[i]))
        left_NN.append((left_NN_i[i],left_NN_j[i]))
        bottom_NN.append((bottom_NN_i[i],bottom_NN_j[i]))
        right_NN.append((right_NN_i[i],right_NN_j[i]))
        
    # convert all tuples into list to make it immutable     
    top_NN = list(top_NN)
    left_NN = list(left_NN)
    bottom_NN = list(bottom_NN)
    right_NN = list(right_NN)   
    
    coords = top_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    top_NN = coords
    
    coords = left_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    left_NN = coords
    
    coords = right_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    right_NN = coords
    
    coords = bottom_NN
    coords = [coord for coord in coords if not any(number < 0 for number in coord)]
    bottom_NN = coords   
    
    coords = top_NN
    top_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            top_NN.append(coord) 
            
    coords = left_NN
    left_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            left_NN.append(coord) 
            
    coords = right_NN
    right_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            right_NN.append(coord) 
            
    coords = bottom_NN
    bottom_NN = []
    for coord in coords:
        if not coord[0] >= img.shape[0] and not coord[1] >= img.shape[1]:
            bottom_NN.append(coord) 
    
    NN_List = top_NN + left_NN + bottom_NN + right_NN    
    NN_List = list( dict.fromkeys(NN_List) )

    return NN_List



class pixel_block():
    def __init__(self, parent=None):
        self.shape_cluster = [] 
        self.cluster_bounds = []
        self.NN_List = []
        self.img = []
        self.img_shape = []
        self.pos_i = 0 
        self.pos_j = 0     
        self.thread_cluster = []
                        
    def get_cluster_iteration(self, img, pos_i, pos_j, new = True):
        
        if img[pos_i, pos_j] != 0:
            return 0
        
        if img[pos_i, pos_j] == 0 and new == True:
            self.img = img
            neighbour = (pos_i, pos_j)
            self.pos_i, self.pos_j = pos_i, pos_j
            self.shape_cluster.append(neighbour)
                 
        possible_pixels = get_nearest_neighbours(img, pos_i, pos_j, 1)

        for pixel in possible_pixels:
            if pixel in self.shape_cluster:
                continue

            if img[pixel[0], pixel[1]] != 0:
                continue

            if img[pixel[0], pixel[1]] == 0:
                self.shape_cluster.append(pixel)
                img[pixel[0], pixel[1]] = 1
                
                neighbour_list = get_nearest_neighbours(img, pixel[0], pixel[1], 1)
                for neighbour in neighbour_list:
                    possible_pixels.append(neighbour)                          
                
        if not self.shape_cluster:
            return 0
        else:
            return 1    
        
    def count_pixel(self, img, pix): 
        count = 0
        for i in range(0, img.shape[0]):
            for j in range(0, img.shape[1]):
                if img[i,j] == pix:
                    count = count + 1
                    
        return count
    
    def repeat_points_area(self, img, check_area):
        point_area_check = False
        if len(self.shape_cluster) < check_area:
            point_area_check = True
            for coord in self.shape_cluster:
                img[coord] = 255
                
        return point_area_check, img
            
    def calc_min_distance(self, pos, max_row, max_col):
        horiz_plus = 1
        horiz_minus = 1
        vert_plus = 1
        vert_minus = 1
        max_thickness = 3
        
        while pos[1] + horiz_plus < max_col + 1:       
            if pos[1] + horiz_plus > self.img.shape[1]:
                break
            
            if self.img[pos[0], pos[1] + horiz_plus] == 255:
                break    
                
            horiz_plus = horiz_plus + 1

        while pos[1] - horiz_minus >= 0:
            if self.img[pos[0], pos[1] - horiz_minus] == 255:
                break   
                
            horiz_minus = horiz_minus + 1
            
        while pos[0] - vert_minus >= 0: 
            if self.img[pos[0] - vert_minus, pos[1]] == 255:
                break  
                
            vert_minus = vert_minus + 1

        while pos[0] + vert_plus < max_row + 1:
            if pos[0] + vert_plus > self.img.shape[0]:
                break
            
            if self.img[pos[0] + vert_plus, pos[1]] == 255:
                break
            
            vert_plus = vert_plus + 1
                 
        vert_thickness = vert_plus + vert_minus
        horiz_thickness = horiz_plus + horiz_minus   
        thickness = min(horiz_thickness, vert_thickness)
        
        thickness_midpoint = False
        if thickness == horiz_thickness:
            if horiz_plus == horiz_minus and horiz_plus>1:
                thickness_midpoint = True
                
        if thickness == vert_thickness:
            if vert_plus == vert_minus and vert_plus > 1:
                thickness_midpoint = True
        
        if thickness < max_thickness:
            if thickness_midpoint is False:
                self.thread_cluster.append(pos)
                i = -1 * max_thickness
                while i < max_thickness:
                    if pos[1] + i < max_col: 
                        if self.img[pos[0], pos[1] + i] == 1:
                            self.thread_cluster.append((pos[0], pos[1] + i))
                    i += 1

                i = -1 * max_thickness
                while i < max_thickness:
                    if pos[0] + i < max_row: 
                        if self.img[pos[0] + i, pos[1]] == 1:
                            self.thread_cluster.append((pos[0] + i, pos[1]))
                    i += 1
    
    def connected_cluster_seggragation(self, img):
        self.thread_cluster = []
        thread_cluster_temp = []
        self.img = img
        max_row = 0
        max_col = 0
        for coord in self.shape_cluster:
            if self.img[coord] == 0:
                self.img[coord] = 1
            
            if coord[0] > max_row:
                max_row = coord[0]
                
            if coord[1] > max_col:
                max_col = coord[1]
        
        for coord in self.shape_cluster:
            self.calc_min_distance(coord, max_row, max_col)
            
        for coord in self.thread_cluster:
            mark_possibility_NN1 = False
            mark_possibility_NN2 = False
            possible_pixels = get_nearest_neighbours(self.img, coord[0], coord[1], 1)           
            for possible_cord in possible_pixels:
                if possible_cord in self.thread_cluster:
                    mark_possibility_NN1 = True
            
            if coord[0] > 0 and coord[1] > 0:
                possible_pixels = get_nearest_neighbours(self.img, coord[0]-1, coord[1]-1, 2)           
                for possible_cord in possible_pixels:
                    if possible_cord in self.thread_cluster:
                        mark_possibility_NN2 = True
                        
            else:
                # Since the pixel is in the edge
                mark_possibility_NN2 = False
                    
            # To check if the thread is atleast 3 pixels long
            if mark_possibility_NN1 is True and mark_possibility_NN2 is True:
                thread_cluster_temp.append(coord)
          
        if thread_cluster_temp:
            for elem in thread_cluster_temp:
                if self.img[elem] == 1:
                    self.img[elem] = 2

        self.thread_cluster = thread_cluster_temp
        return self.img
                
    def view_closed_shape(self):
        max_row = 10
        max_col = 10
        sq_diam = 0
        self.img_shape = []
        print("current cluster cordinate: ", self.pos_i, self.pos_j)
    
        for coord in self.shape_cluster:
            if coord[0] > max_row:
                max_row = coord[0]
                
            if coord[1] > max_col:
                max_col = coord[1]
                
            sq_diam = max(max_row, max_col)
            div = int(sq_diam/5)
            sq_diam = div * 5 + 5
        
        self.img_shape = np.full((sq_diam, sq_diam), 255)
        for i in range(sq_diam):
            self.img_shape[0][i] = 150
            self.img_shape[i][0] = 150
            self.img_shape[i][sq_diam-1] = 150
            self.img_shape[sq_diam-1][i] = 150
                                
        for coord in self.shape_cluster:
            self.img_shape[coord] = 0
        
        if sq_diam > 0:
            view1_image(self.img_shape)
        else:
            print("no cluster detected..")          
            
        if self.thread_cluster:
            self.img_shape = []
            for coord in self.thread_cluster:
                if coord[0] > max_row:
                    max_row = coord[0]

                if coord[1] > max_col:
                    max_col = coord[1]

                sq_diam = max(max_row, max_col)
                div = int(sq_diam/5)
                sq_diam = div * 5 + 5

            self.img_shape = np.full((sq_diam, sq_diam), 255)
            for i in range(sq_diam):
                self.img_shape[0][i] = 150
                self.img_shape[i][0] = 150
                self.img_shape[i][sq_diam-1] = 150
                self.img_shape[sq_diam-1][i] = 150

            for coord in self.thread_cluster:
                self.img_shape[coord] = 0

            if sq_diam > 0:
                view1_image(self.img_shape)
            else:
                print("no thread detected..")
                
    
    def join_broken_threads(self, area = 4):
        # Current cluster in focus is a point        
        if area <= int(len(self.shape_cluster)):
            for coord in self.shape_cluster:
                if coord[0] - 1 >= 0 and coord[1] - 1 >= 0:
                    pos_i = coord[0] - 1
                    pos_j = coord[1] - 1
                    possible_points = get_nearest_neighbours(self.img, pos_i, pos_j, 3)

                    for possible_coord in possible_points:
                        if (self.img[possible_coord] == 1 or self.img[possible_coord] == 0):
                            if possible_coord not in self.shape_cluster:
                                self.img = cv2.line(self.img, (coord[1], coord[0]), (possible_coord[1], possible_coord[0]), (2, 2, 2), thickness = 1, lineType=8)


            for coord in self.shape_cluster:
                if coord[0] - 2 >= 0 and coord[1] - 2 >= 0:
                    pos_i = coord[0] - 2
                    pos_j = coord[1] - 2
                    possible_points = get_nearest_neighbours(self.img, pos_i, pos_j, 5)
                    for possible_coord in possible_points:
                        if (self.img[possible_coord] == 1 or self.img[possible_coord] == 0):
                            if possible_coord not in self.shape_cluster:
                                self.img = cv2.line(self.img, (coord[1], coord[0]), (possible_coord[1], possible_coord[0]), (2, 2, 2), thickness = 1, lineType=8)

        self.img[self.img == 2] = 1
        return self.img 

    
    
    
# To detect clusters, join points to closeby clusters/threads, count the remaining points and eliminate them
def porosity_operation1(img):
    img_trial = np.copy(img) 
    area_points = 0
    val = 0
    image_exist_bit = False
    
    for i in range(0, img_trial.shape[0]):
        for j in range(0, img_trial.shape[1]):
            image_exist_bit = True
            block = pixel_block()
            val = block.get_cluster_iteration(img_trial, i, j, True)
            if val == 1:
                img_trial = block.join_broken_threads(1)
    
    if image_exist_bit:
        area_points = block.count_pixel(img_trial, 0)
    else:
        area_points = 0
        
    img_trial[img_trial == 0] = 255
    img_trial[img_trial == 1] = 0  
    return img_trial
    
    
# operation 2: Recalculate point area by giving in block size as arguement - also cleans the image    
def porosity_operation2(img):
    img_trial = np.copy(img)
    area_points = 0
    val = 0
    for i in range(0, img_trial.shape[0]):
        for j in range(0, img_trial.shape[1]):  
            block = pixel_block()
            val = block.get_cluster_iteration(img_trial, i, j, True)
            if val == 1:
                point_area_check, img_trial2 = block.repeat_points_area(img_trial, 4)
                if point_area_check is True:
                    area_points = area_points + 1
                img_trial = block.join_broken_threads(1)
    img_trial[img_trial == 0] = 255
    img_trial[img_trial == 1] = 0
            
    return img_trial, area_points
    
    
# operation 3: isolate threads from clusters and count them - Connected-component labeling
def porosity_operation3(img):
    img_trial = np.copy(img)
    area_threads = 0
    val = 0
    for i in range(0, img_trial.shape[0]):
        for j in range(0, img_trial.shape[1]):
            block = pixel_block()
            val = block.get_cluster_iteration(img_trial, i, j, True)
            if val == 1:
                img_trial = block.connected_cluster_seggragation(img_trial)

    img_trial[img_trial == 2] = 150
    area_threads = block.count_pixel(img_trial, 150)
    return img_trial, area_threads
    
    
# operation 4: gets the final cluster area
def porosity_operation4(img):
    img_trial = np.copy(img)
    area_clusters = 0
    img_trial[img_trial == 150] = 255
    img_trial[img_trial == 1] = 0
    block = pixel_block()
    area_clusters = block.count_pixel(img_trial, 0)   
    return img_trial, area_clusters    



def df_slice_porosity_update(df_slice):
    df_slice_porosity = pd.DataFrame({'Sample': str,
                                       'Image_ID': str,
                                       'Slice_ID':int,
                                       'Points_area': [],
                                       'Threads_area': [],
                                       'Clusters_area': [],
                                       'Total_area': []})

    Porosity_data_file = os.path.join(path_samples, "Porosity_dataframe.json")
    continue_operation = False
    
    if os.path.isfile(Porosity_data_file): 
        print("Using last saved porosity data..(Delete -Porosity_dataframe.json- to begin NEW operation..)")
        df_slice_porosity = pd.read_json(Porosity_data_file, orient = 'columns') 
        df_slice_porosity.sort_index(inplace=True)
        required_number = df_slice.shape[0]
        current_number = df_slice_porosity.shape[0]
        if required_number > current_number:
            print("Completing the porosity calculations for remaining slices of exisitng file")
            continue_operation = True  
            
        else:
            return df_slice_porosity
                

    else: 
        print("No previously saved porosity data found, beginning NEW slicing operation..")
        max_count = df_slice.shape[0]
        progress_bar = IntProgress(min=0, max=max_count) 
        display(progress_bar) 
        count = 0

        df_slice_porosity_temp = pd.DataFrame({'Sample': str,
                                               'Image_ID': str,
                                               'Slice_ID':int,
                                               'Points_area': [],
                                               'Threads_area': [],
                                               'Clusters_area': [],
                                               'Total_area': []})

        for index, row in df_slice.iterrows():
            Sample = row['Sample']
            Image_ID = row['Image_ID']
            Slice_ID = row['Slice_ID']
            Slice_Array_Threshold = row['Slice_Array_Threshold']
            
            print("[0]Processing Image_ID: ", Image_ID," with Slice ID: ", Slice_ID)
            Slice_Array_Threshold_img = np.asarray(Slice_Array_Threshold, dtype='uint8')
            img_op1 = porosity_operation1(Slice_Array_Threshold_img)
            img_op2, area_points = porosity_operation2(img_op1)
            img_op3, area_threads = porosity_operation3(img_op2)
            img_op4, area_clusters = porosity_operation4(img_op3)
            area_total = Slice_Array_Threshold_img.shape[0] * Slice_Array_Threshold_img.shape[1]

            print("[1]Porosity Detection of Image_ID: ", Image_ID, " with Slice_ID: ", Slice_ID, " Complete ")
            progress_bar.value += 1 
            time.sleep(.1)
            
            df_slice_porosity_temp['Sample'] = pd.Series(Sample)
            df_slice_porosity_temp['Image_ID'] = pd.Series(Image_ID)
            df_slice_porosity_temp['Slice_ID'] = pd.Series(Slice_ID)
            df_slice_porosity_temp['Points_area'] = pd.Series(area_points)
            df_slice_porosity_temp['Threads_area'] = pd.Series(area_threads)
            df_slice_porosity_temp['Clusters_area'] = pd.Series(area_clusters)
            df_slice_porosity_temp['Total_area'] = pd.Series(area_total)

            df_slice_porosity = df_slice_porosity.append(df_slice_porosity_temp, ignore_index = True)
            df_slice_porosity.to_json(Porosity_data_file, orient='columns')
            
        return df_slice_porosity


    if continue_operation is True:
        max_count = df_slice.shape[0]
        progress_bar = IntProgress(min=0, max=max_count) 
        display(progress_bar) 
        count = 0
        display(df_slice_porosity.tail())   
        df_slice_porosity_temp = pd.DataFrame({'Sample': str,
                                               'Image_ID': str,
                                               'Slice_ID':int,
                                               'Points_area': [],
                                               'Threads_area': [],
                                               'Clusters_area': [],
                                               'Total_area': []})

        for index, row in df_slice.iterrows():
            Sample = row['Sample']
            Image_ID = row['Image_ID']
            Slice_ID = row['Slice_ID']
            Slice_Array_Threshold = row['Slice_Array_Threshold']

            if index < current_number:
                progress_bar.value += 1 
                time.sleep(.1)
                pass

            else:
                print("[0]Processing Image_ID: ", Image_ID," with Slice ID: ", Slice_ID)
                Slice_Array_Threshold_img = np.asarray(Slice_Array_Threshold, dtype='uint8')
                img_op1 = porosity_operation1(Slice_Array_Threshold_img)
                img_op2, area_points = porosity_operation2(img_op1)
                img_op3, area_threads = porosity_operation3(img_op2)
                img_op4, area_clusters = porosity_operation4(img_op3)
                area_total = Slice_Array_Threshold_img.shape[0] * Slice_Array_Threshold_img.shape[1]
                
                print("[1]Porosity Detection of Image_ID: ", Image_ID, " with Slice_ID: ", Slice_ID, " Complete ")
                progress_bar.value += 1 
                time.sleep(.1)
                df_slice_porosity_temp['Sample'] = pd.Series(Sample)
                df_slice_porosity_temp['Image_ID'] = pd.Series(Image_ID)
                df_slice_porosity_temp['Slice_ID'] = pd.Series(Slice_ID)
                df_slice_porosity_temp['Points_area'] = pd.Series(area_points)
                df_slice_porosity_temp['Threads_area'] = pd.Series(area_threads)
                df_slice_porosity_temp['Clusters_area'] = pd.Series(area_clusters)
                df_slice_porosity_temp['Total_area'] = pd.Series(area_total)
                
                df_slice_porosity = df_slice_porosity.append(df_slice_porosity_temp, ignore_index = True)
                df_slice_porosity.to_json(Porosity_data_file, orient='columns')
                
        return df_slice_porosity
    
    
def df_slice_porosity_file_display():
    Porosity_data_file = os.path.join(path_samples, "Porosity_dataframe.json")
    if os.path.isfile(Porosity_data_file): 
        print("Loading the latest (Temporary/Full) slice porosity data..")
        df_slice_porosity_new = pd.read_json(Porosity_data_file, orient = 'columns') 
        df_slice_porosity_new.sort_index(inplace=True)
        print("The slice file at: ", Porosity_data_file, " has the following last rows: ")
        display(df_slice_porosity_new.tail(3)) 
    return df_slice_porosity_new


# def df_mercury_data_extract():
#     Hg_dict = {}
#     for root, dirs, files in os.walk(path_samples):
#         for name in files:
#             filename = os.path.join(root, name)
#             if os.path.basename(filename) == "Hg_Porosimetry.xlsx":
#                 key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
#                 Hg_dict[key] = filename

#     start_row = 0
#     Hg_df_dict = {}
#     datas_ = []
#     for key in Hg_dict:
#         xl_workbook = xlrd.open_workbook(Hg_dict[key])
#         sheet_names = xl_workbook.sheet_names()
#         xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
#         num_cols = xl_sheet.ncols   # Number of columns
#         num_rows = xl_sheet.nrows
#         for row_idx in range(0, num_rows):    # Iterate through rows
#             row = row_idx
#             for col_idx in range(0, num_cols):  # Iterate through columns
#                 cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
#                 col = col_idx
#                 cell_obj_trunc = str(cell_obj.value).replace(" ", "")
#                 if cell_obj_trunc == "DATAREPORT":
#                     start_row = row + 1

#         excel_file = Hg_dict[key]
#         df = pd.read_excel(excel_file)
#         df_sheet1 = pd.read_excel(excel_file, sheet_name = 0, skiprows = start_row)
#         df_sheet1 = df_sheet1.drop([0])
#         df_sheet1.reset_index(inplace=True)
#         df_sheet1 = df_sheet1.drop(columns=['index'])
#         df_sheet1 = df_sheet1.loc[:, ['Radius', 'P[%]']]
#         df_sheet1.dropna(inplace=True)
#         datas_.append((key, df_sheet1['Radius'].values, df_sheet1['P[%]'].values))
#         Hg_df_dict[key] = df_sheet1
        
#     return datas_, Hg_df_dict



# def df_mercury_common_cordinates(datas_, Hg_df_dict, df_slice_porosity, num = 10, kind = "linear"):
    
#     def reinterpolate(x_old, y_old, x_new, kind = "linear"):
#         f = interp1d(x_old, y_old, kind = kind)
#         return f(x_new)

#     Hg_porosity_df = pd.DataFrame({'Sample': str,
#                        'Image_ID': str,
#                        'Slice_ID':int,
#                        'Radii': [],
#                        'Ps': [],
#                        'log_Radii': [],
#                        'log_Radii_new': [],
#                        'Ps_new': []})

#     Hg_porosity_df_temp = pd.DataFrame({'Sample': str,
#                        'Image_ID': str,
#                        'Slice_ID':int,
#                        'Radii': [],
#                        'Ps': [],
#                        'log_Radii': [],
#                        'log_Radii_new': [],
#                        'Ps_new': []})
    
#     porosity_df = pd.DataFrame(datas_, columns=['Sample',"Radii","Ps"])
#     porosity_df["Radii"] = porosity_df["Radii"].map(lambda array: np.array(array).astype(float))
#     porosity_df["log_Radii"] = porosity_df["Radii"].map(np.log)
#     rlogmin = porosity_df["log_Radii"].map(min).max()
#     rlogmax  =porosity_df["log_Radii"].map(max).min()
#     log_Radii_new = np.linspace(rlogmin, rlogmax, num = num)

    
#     for index, row in df_slice_porosity.iterrows():
#         Sample = row['Sample']
#         Image_ID = row['Image_ID']
#         Slice_ID = row['Slice_ID']
#         Hg_df = Hg_df_dict[Sample]
#         Radii = Hg_df['Radius'].tolist()
#         log_Radii = np.log(Radii)
#         Ps = Hg_df['P[%]'].tolist()
#         Ps_new = reinterpolate(log_Radii, Ps, log_Radii_new, kind = kind)

#         Hg_porosity_df_temp['Sample'] = pd.Series(Sample)
#         Hg_porosity_df_temp['Image_ID'] = pd.Series(Image_ID)
#         Hg_porosity_df_temp['Slice_ID'] = pd.Series(Slice_ID)
#         Hg_porosity_df_temp['Radii'] = pd.Series([Radii])
#         Hg_porosity_df_temp['Ps'] = pd.Series([Ps])
#         Hg_porosity_df_temp['log_Radii'] = pd.Series([log_Radii])
#         Hg_porosity_df_temp['log_Radii_new'] = pd.Series([log_Radii_new])
#         Hg_porosity_df_temp['Ps_new'] = pd.Series([Ps_new])
        
#         Hg_porosity_df = Hg_porosity_df.append(Hg_porosity_df_temp, ignore_index = True)
        
#     return Hg_porosity_df


# def df_parameters_data_extract():
#     parameters_filename_dict = {}
#     for root, dirs, files in os.walk(path_samples):
#         for name in files:
#             filename = os.path.join(root, name)
#             if os.path.basename(filename) == "parameters.xlsx":
#                 key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
#                 parameters_filename_dict[key] = filename
    
#     row = 0
#     parameters_list_dict = {}
#     for key in parameters_filename_dict:
#         parameters_dict = {}
#         xl_workbook = xlrd.open_workbook(parameters_filename_dict[key])
#         sheet_names = xl_workbook.sheet_names()
#         xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
#         num_cols = xl_sheet.ncols   # Number of columns
#         num_rows = xl_sheet.nrows
        
#         for col_idx in range(0, num_cols):
#             cell_key = str(xl_sheet.cell(0, col_idx).value).replace(" ", "")
#             cell_value = str(xl_sheet.cell(1, col_idx).value).replace(" ", "")
#             parameters_dict[cell_key] = cell_value  
            
#         parameters_list_dict[row] = parameters_dict
#         row = row + 1
        
#     df_parameters = pd.DataFrame.from_dict(parameters_list_dict,orient='index')
#     return df_parameters


# def df_parameters_match(df_parameters, df_slice_porosity, bounded, max_powder_dia, div):
#     df_parameters_copy = df_parameters.copy()
#     if bounded is True:
        
#         df_parameters_copy['Powder_Diameter_0.1_(microns)'] = (df_parameters['Powder_Diameter_0.1_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
#         df_parameters_copy['Powder_Diameter_0.5_(microns)'] = (df_parameters['Powder_Diameter_0.5_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
#         df_parameters_copy['Powder_Diameter_0.9_(microns)'] = (df_parameters['Powder_Diameter_0.9_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
        
#     df_slice_parameters = df_slice_porosity[['Sample', 'Image_ID']].merge(df_parameters_copy, left_on='Sample', right_on='Name')
#     df_slice_parameters = df_slice_parameters.drop(['Name'], axis=1)
    
    
#     return df_slice_parameters

def df_roughness_data_extract():
    df_roughness = pd.DataFrame({'Sample': str,
                       'Roughness_Type': [],
                       'Lambda_c_(microns)': [],
                       'Ra_(microns)': [],
                       'Rq_(microns)': [],
                       'Rz_(microns)': []})
    
    sample = []
    roughness_type = []
    lambda_c = []
    ra = []
    rq = []
    rz = []
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "roughness.txt":
                sample_name = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                sample_type = os.path.basename(os.path.dirname(filename))
                sample.append(sample_name)
                roughness_type.append(sample_type)
                with open(filename, "r") as f:
                    data = f.readlines()
                    for index, line in enumerate(data):
                        if index == 1:
                            words = line.split(";")
                            for index, value in enumerate(words):
                                if index == 4:
                                    lambda_c.append(value)   
                                if index == 6:
                                    ra.append(value)
                                if index == 9:
                                    rq.append(value)
                                if index == 12:
                                    rz.append(value)
     
    df_roughness['Sample'] = pd.Series(sample)
    df_roughness['Roughness_Type'] = pd.Series(roughness_type)
    df_roughness['Lambda_c_(microns)'] = pd.Series(lambda_c)
    df_roughness['Ra_(microns)'] = pd.Series(ra)
    df_roughness['Rq_(microns)'] = pd.Series(rq)
    df_roughness['Rz_(microns)'] = pd.Series(rz)

    return df_roughness



def img_to_slices(img, diam = 224, no_of_slices = 20): 
    # 1st Operation
    def image_dilate(img):
        img_white_matrix = np.copy(img)
        blur = cv2.GaussianBlur(img_white_matrix, (5, 5), 0)
        ret, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        morph = th.copy()
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
        morph = cv2.morphologyEx(morph, cv2.MORPH_CLOSE, kernel)
        morph = cv2.morphologyEx(morph, cv2.MORPH_OPEN, kernel)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        erosion = cv2.erode(img_white_matrix, kernel, iterations = 1)
        dilate = cv2.dilate(erosion, kernel, iterations = 2)
        erosion = cv2.erode(dilate, kernel, iterations = 1)
        img_white_matrix_dilate = cv2.dilate(erosion, kernel, iterations = 1) 

        return img_white_matrix_dilate

    # 2nd Operation
    def delete_non_matrix(img_original):
        img = np.copy(img_original)
        neighbours = np.zeros((4,4))
        rows = img.shape[0]
        cols = img.shape[1]
        inter = []
        inter_mean = 0
        for j in range(0, cols):
            for i in range(0, rows):
                if img[i, j] == 255:
                    inter = img[0:i + 1, j:j + 1].ravel()
                    inter_mean = np.mean(inter[0:i])       
                    if inter_mean < 2.0:
                        img[0:i + 1, j:j + 1] = 120
                    break

        for j in range(0, cols):
            for i in range(rows-1, 0, -1):
                if img[i, j] == 255:
                    inter = img[i + 1:rows, j:j + 1].ravel()
                    inter_mean = np.mean(inter[0:i])    
                    if inter_mean < 2.0 and not math.isnan(inter_mean):
                        img[i + 1:rows, j:j + 1] = 120
                    break         
        return img

    # 3rd Operation
    def find_matrix_threshhold(img_part):
        naux = 0
        naux_color = 0 
        threshold = 0
        threshold_offset = 20
        threshold_fine = []
        ascending_pixel = []
        curr_max_pixel = 0
        curr_max_pixel_index = 0
        pix_width = img_part.shape[0]
        pix_height = img_part.shape[1]

        for i in range(256):
            threshold_fine.append(0)
            ascending_pixel.append(0)

        for i in range(pix_width):
            for j in range(pix_height):
                pix = int(img_part[i,j])
                threshold_fine[pix] = threshold_fine[pix] + 1

            if naux <= threshold_fine[pix]: 
                naux = threshold_fine[pix] 
                naux_color = pix

        for i in range(len(threshold_fine)):
            curr_max_pixel = np.max(threshold_fine)
            curr_max_pixel_index_tuple = np.where(threshold_fine == curr_max_pixel)
            curr_max_pixel_index = curr_max_pixel_index_tuple[0][0]  
            ascending_pixel[i] = curr_max_pixel_index
            threshold_fine[curr_max_pixel_index] = 0

        for i in range(len(ascending_pixel)):
            if ascending_pixel[i] > 150:
                threshold = ascending_pixel[i] - threshold_offset
                break
            else:
                threshold = 150      

        return threshold

    # 4th Operation
    def matrix_find_thresh(img):
        img_copy = np.copy(img)
        height_image = img_copy.shape[0]
        width_image = img_copy.shape[1]
        div = 5     

        height_part = int(height_image/div)
        width_part = int(width_image/div)

        dif_height = height_image - int(height_part*div)
        dif_width = width_image - int(width_part*div)

        i = 0
        while i<height_image:
            j = 0
            while j<width_image:
                if (i + height_part <= height_image) and (j + width_part <= width_image):

                    img_part = img_copy[i:i+height_part, j:j+width_part]
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[i:i+height_part, j:j+width_part] = th

                if j + width_part > width_image:
                    img_part = img_copy[0:height_part, j:j+dif_width] 
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[0:height_part, j:j+dif_width]  = th

                if i + height_part > height_image:
                    img_part = img_copy[i:i+dif_height, 0:dif_width] 
                    thresh_img_part = find_matrix_threshhold(img_part)
                    ret,th = cv2.threshold(img_part, thresh_img_part, 255, cv2.THRESH_BINARY)
                    img_copy[i:i+dif_height, 0:dif_width] = th

                j = j + width_part
            i = i + height_part

        return img_copy
    
    bottom = img.shape[0]
    right_side = img.shape[1]
    img_white_matrix_dilate = image_dilate(img)
    img_white_matrix_dilate_without_matrix = delete_non_matrix(img_white_matrix_dilate)
    img_after_find_matrix = matrix_find_thresh(img_white_matrix_dilate_without_matrix)
    slices_main_thresh = img_after_find_matrix
    
    slice_centres_x = 0
    slice_centres_y = 0
    slice_sq_dim = diam  
    slices_thresh = [np.zeros((slice_sq_dim, slice_sq_dim)) for x in range(no_of_slices)]
    
    for sl_no in range(no_of_slices):
        slice_centres_x = (np.random.randint(slice_sq_dim, right_side, size=1))[0]
        slice_centres_y = (np.random.randint(slice_sq_dim, bottom - slice_sq_dim, size=1))[0]
        
        x1 = slice_centres_x - int( slice_sq_dim / 2 )
        y1 = slice_centres_y - int( slice_sq_dim / 2 ) 
        x2 = slice_centres_x + int( slice_sq_dim / 2 )
        y2 = slice_centres_y + int( slice_sq_dim / 2 )
        
        while y2 > bottom:
            y1 = y1 - (np.random.randint(1, 100, size=1))[0]
            y2 = y1 + slice_sq_dim
                
        while x2 > right_side:
            x1 = x1 - (np.random.randint(1, 100, size=1))[0]
            x2 = x1 + slice_sq_dim
            
        slices_thresh[sl_no] = slices_main_thresh[y1:y2, x1:x2]
        
    return slices_thresh

def df_mercury_data_extract():
    Hg_dict = {}
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "Hg_Porosimetry.xlsx":
                key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                Hg_dict[key] = filename

    start_row = 0
    Hg_df_dict = {}
    datas_ = []
    for key in Hg_dict:
        xl_workbook = xlrd.open_workbook(Hg_dict[key])
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        num_cols = xl_sheet.ncols   # Number of columns
        num_rows = xl_sheet.nrows
        for row_idx in range(0, num_rows):    # Iterate through rows
            row = row_idx
            for col_idx in range(0, num_cols):  # Iterate through columns
                cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
                col = col_idx
                cell_obj_trunc = str(cell_obj.value).replace(" ", "")
                if cell_obj_trunc == "DATAREPORT":
                    start_row = row + 1

        excel_file = Hg_dict[key]
        df = pd.read_excel(excel_file)
        df_sheet1 = pd.read_excel(excel_file, sheet_name = 0, skiprows = start_row)
        df_sheet1 = df_sheet1.drop([0])
        df_sheet1.reset_index(inplace=True)
        df_sheet1 = df_sheet1.drop(columns=['index'])
        df_sheet1 = df_sheet1.loc[:, ['Radius', 'P[%]']]
        df_sheet1.dropna(inplace=True)
        datas_.append((key, df_sheet1['Radius'].values, df_sheet1['P[%]'].values))
        Hg_df_dict[key] = df_sheet1
        
    return datas_, Hg_df_dict

def df_mercury_common_cordinates(datas_, Hg_df_dict, df_slice_porosity, num = 10, kind = "linear"):
    
    def reinterpolate(x_old, y_old, x_new, kind = "linear"):
        f = interp1d(x_old, y_old, kind = kind)
        return f(x_new)

    Hg_porosity_df = pd.DataFrame({'Sample': str,
                       'Image_ID': str,
                       'Slice_ID':int,
                       'Radii': [],
                       'Ps': [],
                       'log_Radii': [],
                       'log_Radii_new': [],
                       'Ps_new': []})

    Hg_porosity_df_temp = pd.DataFrame({'Sample': str,
                       'Image_ID': str,
                       'Slice_ID':int,
                       'Radii': [],
                       'Ps': [],
                       'log_Radii': [],
                       'log_Radii_new': [],
                       'Ps_new': []})
    
    porosity_df = pd.DataFrame(datas_, columns=['Sample',"Radii","Ps"])
    porosity_df["Radii"] = porosity_df["Radii"].map(lambda array: np.array(array).astype(float))
    porosity_df["log_Radii"] = porosity_df["Radii"].map(np.log)
    rlogmin = porosity_df["log_Radii"].map(min).max()
    rlogmax  =porosity_df["log_Radii"].map(max).min()
    log_Radii_new = np.linspace(rlogmin, rlogmax, num = num)

    for index, row in df_slice_porosity.iterrows():
        Sample = row['Sample']
        Image_ID = row['Image_ID']
        Slice_ID = row['Slice_ID']
        Hg_df = Hg_df_dict[Sample]
        Radii = Hg_df['Radius'].tolist()
        log_Radii = np.log(Radii)
        Ps = Hg_df['P[%]'].tolist()
        Ps_new = reinterpolate(log_Radii, Ps, log_Radii_new, kind = kind)

        Hg_porosity_df_temp['Sample'] = pd.Series(Sample)
        Hg_porosity_df_temp['Image_ID'] = pd.Series(Image_ID)
        Hg_porosity_df_temp['Slice_ID'] = pd.Series(Slice_ID)
        Hg_porosity_df_temp['Radii'] = pd.Series([Radii])
        Hg_porosity_df_temp['Ps'] = pd.Series([Ps])
        Hg_porosity_df_temp['log_Radii'] = pd.Series([log_Radii])
        Hg_porosity_df_temp['log_Radii_new'] = pd.Series([log_Radii_new])
        Hg_porosity_df_temp['Ps_new'] = pd.Series([Ps_new])
        
        Hg_porosity_df = Hg_porosity_df.append(Hg_porosity_df_temp, ignore_index = True)
        
    return Hg_porosity_df

def df_parameters_data_extract():
    print("hi")
    parameters_filename_dict = {}
    for root, dirs, files in os.walk(path_samples):
        for name in files:
            filename = os.path.join(root, name)
            if os.path.basename(filename) == "parameters.xlsx":
                key = os.path.basename(os.path.dirname(os.path.dirname(filename)))
                parameters_filename_dict[key] = filename
    
    row = 0
    parameters_list_dict = {}
    for key in parameters_filename_dict:
        parameters_dict = {}
        xl_workbook = xlrd.open_workbook(parameters_filename_dict[key])
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        num_cols = xl_sheet.ncols   # Number of columns
        num_rows = xl_sheet.nrows
        
        for col_idx in range(0, num_cols):
            cell_key = str(xl_sheet.cell(0, col_idx).value).replace(" ", "")
            cell_value = str(xl_sheet.cell(1, col_idx).value).replace(" ", "")
            parameters_dict[cell_key] = cell_value  
            
        parameters_list_dict[row] = parameters_dict
        row = row + 1
        
    df_parameters = pd.DataFrame.from_dict(parameters_list_dict,orient='index')
    return df_parameters

def df_parameters_match(df_parameters, df_slice_porosity, bounded, max_powder_dia, div):
    df_parameters_copy = df_parameters.copy()
    if bounded is True:
        df_parameters_copy['Powder_Diameter_0.1_(microns)'] = (df_parameters['Powder_Diameter_0.1_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
        df_parameters_copy['Powder_Diameter_0.5_(microns)'] = (df_parameters['Powder_Diameter_0.5_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
        df_parameters_copy['Powder_Diameter_0.9_(microns)'] = (df_parameters['Powder_Diameter_0.9_(microns)'].astype(float)/max_powder_dia * div).astype(int) + 1
        
    df_slice_parameters = df_slice_porosity[['Sample', 'Image_ID']].merge(df_parameters_copy, left_on='Sample', right_on='Name')
    df_slice_parameters = df_slice_parameters.drop(['Name'], axis=1)
    
    
    return df_slice_parameters


